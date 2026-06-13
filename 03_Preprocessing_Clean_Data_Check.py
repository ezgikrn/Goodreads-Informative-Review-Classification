from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# ==================== 1. DOSYA VE BEKLENEN KOLONLAR ====================
BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "MEB_100_Goodreads_ML_LLM.xlsx"

REQUIRED_SHEETS = ["Yorumlar_On_Isleme", "Puan_Analizi", "Informative_ML_Dataset"]

REQUIRED_COLUMNS = {
    "Yorumlar_On_Isleme": [
        "MEB eser", "Goodreads eser", "Puan", "yorum_raw", "yorum_on_isleme",
        "kelime_sayisi", "yorum_uzunlugu", "cikarilacak_mi", "cikarma_nedeni", "Review ID",
    ],
    "Puan_Analizi": [
        "MEB eser", "Goodreads eser", "Puan", "yorum_temiz", "model_metni_lemma",
        "kelime_sayisi", "yorum_uzunlugu", "final_model_token_sayisi", "Review ID",
    ],
    "Informative_ML_Dataset": [
        "orneklem_id", "Review ID", "yorum_raw", "yorum_temiz", "manuel_informative",
        "Puan", "kelime_sayisi", "yorum_uzunlugu", "split", "cv_fold(10)",
    ],
}


# ==================== 2. KONTROL YARDIMCILARI ====================
def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }


def count_nonempty(sheet, column_name: str) -> int:
    columns = header_map(sheet)
    column = columns[column_name]
    return sum(1 for row in range(2, sheet.max_row + 1) if cell_text(sheet.cell(row, column).value))


def split_counts(sheet) -> Counter:
    columns = header_map(sheet)
    column = columns["split"]
    return Counter(cell_text(sheet.cell(row, column).value) for row in range(2, sheet.max_row + 1))


# ==================== 3. ANA KONTROL ====================
def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    missing_sheets = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError(f"Eksik sayfalar: {missing_sheets}")

    print("Preprocessing kontrolu")
    for sheet_name in REQUIRED_SHEETS:
        sheet = workbook[sheet_name]
        columns = header_map(sheet)
        missing_columns = [column for column in REQUIRED_COLUMNS[sheet_name] if column not in columns]
        if missing_columns:
            raise ValueError(f"{sheet_name} eksik kolonlar: {missing_columns}")
        print(f"{sheet_name}: rows={sheet.max_row - 1}, cols={sheet.max_column}")

    puan_sheet = workbook["Puan_Analizi"]
    ml_sheet = workbook["Informative_ML_Dataset"]
    print(f"Puan_Analizi yorum_temiz dolu: {count_nonempty(puan_sheet, 'yorum_temiz')}/{puan_sheet.max_row - 1}")
    print(f"Informative_ML_Dataset yorum_temiz dolu: {count_nonempty(ml_sheet, 'yorum_temiz')}/{ml_sheet.max_row - 1}")
    print("Informative split dagilimi:", dict(split_counts(ml_sheet)))
    workbook.close()
    print("Kontrol OK: temiz metin ve ML dataset kolonlari hazir.")


if __name__ == "__main__":
    main()
