from __future__ import annotations

import argparse
import json
import os
import re
import time
import urllib.error
import urllib.request
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ==================== 1. DOSYA VE API AYARLARI ====================
BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "MEB_100_Goodreads_ML_LLM.xlsx"
DATASET_SHEET = "Informative_ML_Dataset"
PROGRESS_DIR = BASE_DIR / "LLM_PROGRESS_V3"
API_URL = "https://api.groq.com/openai/v1/chat/completions"
MISTRAL_API_URL = "https://api.mistral.ai/v1/chat/completions"
GEMINI_API_URL_TEMPLATE = "https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
TEST_SPLIT = "test"
LEGACY_TEST_SPLIT = "final" + "_test"
TEST_SPLIT_ALIASES = {TEST_SPLIT, LEGACY_TEST_SPLIT}

DEFAULT_MODELS = {
    "70b": "llama-3.3-70b-versatile",
    "8b": "llama-3.1-8b-instant",
    "qwen": "qwen/qwen3-32b",
    "gptoss120b": "openai/gpt-oss-120b",
    "mistral": "mistral-medium-3-5",
    "gemma": "gemma-4-31b-it",
}

LLM_SHEETS = [
    ("LLM_Groq_Llama70B", "Groq", "Llama 70B"),
    ("LLM_Gemini_Gemma4_31B", "Gemini API", "Gemma 4 31B"),
    ("LLM_Groq_Qwen", "Groq", "Qwen 32B"),
    ("LLM_Mistral_Medium35", "Mistral API", "Mistral Medium 3.5"),
    ("LLM_Groq_GPTOSS120B", "Groq", "GPT-OSS 120B"),
    ("LLM_Groq_Llama8B", "Groq", "Llama 8B"),
]


# ==================== 2. PROMPT ====================
SYSTEM_PROMPT = (
    "Sen Goodreads Turkce kitap yorumlarini bilgilendirici veya bilgilendirici degil olarak "
    "degerlendiren dikkatli bir uzmansin. "
    "Sadece verilen yorum metnine gore karar ver. "
    "Kitap, yazar veya eser hakkinda dis bilgi kullanma. "
    "Cevabini yalnizca gecerli JSON formatinda dondur; aciklama, yorum veya ek metin yazma."
)

ZERO_SHOT_GUIDE = (
    "Gorev: Verilen yorumu bilgilendirici olup olmamasina gore degerlendir.\n\n"
    "Etiketler:\n"
    "1 = Bilgilendirici\n"
    "0 = Bilgilendirici degil\n\n"
    "Bilgilendirici yorum; eser hakkinda karakter, olay/kurgu, tema, anlatim, uslup, "
    "dil/ceviri, icerik veya eserin baglami hakkinda bilgi veren yorumdur.\n\n"
    "Cikti yalnizca gecerli JSON olsun. "
    "JSON icinde sadece label ve confidence alanlari bulunsun. "
    "label degeri yalnizca 0 veya 1 olabilir. "
    "confidence degeri 0.00 ile 1.00 arasinda olabilir.\n\n"
    "Ornek cikti:\n"
    "{\"label\": 1, \"confidence\": 0.87}"
)


# ==================== 3. KOMUT SATIRI ====================
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Informative LLM zero-shot tekrar deneyi.")
    parser.add_argument("--workbook", type=Path, default=WORKBOOK_PATH)
    parser.add_argument("--model-key", choices=["70b", "8b", "qwen", "gptoss120b", "mistral", "gemma"], default="70b")
    parser.add_argument("--model", default=None)
    parser.add_argument("--output-sheet", default=None)
    parser.add_argument("--scope", choices=["test"], default="test")
    parser.add_argument("--repeat", type=int, choices=[1, 2, 3], default=1)
    parser.add_argument("--max-new-records", type=int, default=None)
    parser.add_argument("--sleep", type=float, default=0.25)
    parser.add_argument("--resume", action="store_true", default=True)
    parser.add_argument("--clear-progress", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--update-comparison-only", action="store_true")
    args = parser.parse_args()
    args.scope = normalize_split_name(args.scope)
    return args


# ==================== 4. ENV VE VERI OKUMA ====================
def env_paths() -> list[Path]:
    candidates = [BASE_DIR / ".env"]
    custom_env = os.getenv("GOODREADS_ENV_PATH")
    if custom_env:
        custom_path = Path(custom_env)
        if custom_path not in candidates:
            candidates.append(custom_path)
    return candidates


def load_env_files() -> None:
    for env_path in env_paths():
        if not env_path.exists():
            continue
        for line in env_path.read_text(encoding="utf-8").splitlines():
            stripped = line.strip()
            if not stripped or stripped.startswith("#") or "=" not in stripped:
                continue
            key, value = stripped.split("=", 1)
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = value


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_split_name(value: Any) -> str:
    split = cell_text(value)
    return TEST_SPLIT if split in TEST_SPLIT_ALIASES else split


def is_test_split(value: Any) -> bool:
    return normalize_split_name(value) == TEST_SPLIT


def read_dataset(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[DATASET_SHEET]
    headers = [cell.value for cell in sheet[1]]
    required = [
        "orneklem_id", "Review ID", "MEB eser", "Goodreads eser", "yorum_raw",
        "yorum_temiz", "manuel_informative", "split", "cv_fold(10)",
    ]
    missing = [column for column in required if column not in headers]
    if missing:
        raise ValueError(f"Eksik kolonlar: {missing}")

    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({header: values[index] if index < len(values) else None for index, header in enumerate(headers)})
    workbook.close()
    return rows


def normalize_label(value: Any) -> int:
    label = cell_text(value)
    if label not in {"0", "1"}:
        raise ValueError(f"Gecersiz etiket: {value!r}")
    return int(label)


def test_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in rows if is_test_split(row.get("split"))]


# ==================== 5. ZERO-SHOT PROMPT VE API CAGRI ====================
def build_prompt(row: dict[str, Any]) -> str:
    text = cell_text(row.get("yorum_raw"))
    if len(text) > 1800:
        text = text[:1800] + "..."
    return (
        "ONEMLI: Analiz, gerekce, maddeleme veya prompt tekrari yazma. "
        "Cevap tek satir JSON olsun.\n\n"
        f"{ZERO_SHOT_GUIDE}\n\n"
        "Yorum:\n"
        f"{text}\n\n"
        "Sadece su bicimde tek JSON dondur: {\"label\": 1, \"confidence\": 0.87}"
    )


def call_groq(api_key: str, model: str, prompt: str, sleep_seconds: float, api_url: str = API_URL) -> str:
    is_qwen = "qwen" in model.lower()
    is_gpt_oss = "gpt-oss" in model.lower()
    is_mistral = api_url == MISTRAL_API_URL
    if is_qwen:
        prompt = "/no_think\n" + prompt
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0,
        "max_tokens": 512 if is_gpt_oss else 80 if is_mistral else 40,
    }
    if is_mistral:
        payload["response_format"] = {"type": "json_object"}
    if is_gpt_oss:
        payload["reasoning_effort"] = "low"
    if is_qwen:
        payload["reasoning_effort"] = "none"
        payload["reasoning_format"] = "hidden"

    request = urllib.request.Request(
        api_url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "User-Agent": "Goodreads-Informative-LLM-V3/1.0",
        },
        method="POST",
    )
    for attempt in range(1, 6):
        try:
            with urllib.request.urlopen(request, timeout=90) as response:
                parsed = json.loads(response.read().decode("utf-8"))
                time.sleep(sleep_seconds)
                content = parsed["choices"][0]["message"].get("content", "")
                if isinstance(content, list):
                    content = "".join(str(part.get("text", part)) if isinstance(part, dict) else str(part) for part in content)
                return content or ""
        except urllib.error.HTTPError as error:
            body = error.read().decode("utf-8", errors="replace")
            if error.code in {429, 500, 502, 503, 504} and attempt < 5:
                wait = min(60, 2 ** attempt)
                print(f"API bekleme: HTTP {error.code}, {wait} sn sonra tekrar denenecek.")
                time.sleep(wait)
                continue
            raise RuntimeError(f"API hatasi HTTP {error.code}: {body}") from error
        except urllib.error.URLError as error:
            if attempt < 5:
                wait = min(60, 2 ** attempt)
                print(f"Baglanti bekleme: {wait} sn sonra tekrar denenecek.")
                time.sleep(wait)
                continue
            raise RuntimeError(f"API baglanti hatasi: {error}") from error
    raise RuntimeError("API cagrisi basarisiz.")


def call_gemini(api_key: str, model: str, prompt: str, sleep_seconds: float) -> str:
    payload = {
        "systemInstruction": {"parts": [{"text": SYSTEM_PROMPT}]},
        "contents": [{"role": "user", "parts": [{"text": prompt}]}],
        "generationConfig": {
            "temperature": 0,
            "maxOutputTokens": 512,
            "responseMimeType": "application/json",
            "responseSchema": {
                "type": "OBJECT",
                "properties": {
                    "label": {"type": "INTEGER", "description": "0 or 1"},
                    "confidence": {"type": "NUMBER", "description": "0.00 to 1.00"},
                },
                "required": ["label", "confidence"],
            },
        },
    }
    request = urllib.request.Request(
        GEMINI_API_URL_TEMPLATE.format(model=model, api_key=api_key),
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    for attempt in range(1, 6):
        try:
            with urllib.request.urlopen(request, timeout=90) as response:
                parsed = json.loads(response.read().decode("utf-8"))
                time.sleep(sleep_seconds)
                parts = parsed["candidates"][0]["content"]["parts"]
                text = "".join(part.get("text", "") for part in parts)
                if text.strip():
                    return text
                if attempt < 5:
                    wait = min(60, 2 ** attempt)
                    print(f"Gemini bos cikti verdi, {wait} sn sonra tekrar denenecek.")
                    time.sleep(wait)
                    continue
                raise RuntimeError("Gemini bos cikti verdi.")
        except urllib.error.HTTPError as error:
            body = error.read().decode("utf-8", errors="replace")
            if error.code in {429, 500, 502, 503, 504} and attempt < 5:
                wait = min(60, 2 ** attempt)
                print(f"Gemini bekleme: HTTP {error.code}, {wait} sn sonra tekrar denenecek.")
                time.sleep(wait)
                continue
            raise RuntimeError(f"Gemini API hatasi HTTP {error.code}: {body}") from error
        except urllib.error.URLError as error:
            if attempt < 5:
                wait = min(60, 2 ** attempt)
                print(f"Gemini baglanti bekleme: {wait} sn sonra tekrar denenecek.")
                time.sleep(wait)
                continue
            raise RuntimeError(f"Gemini baglanti hatasi: {error}") from error
    raise RuntimeError("Gemini API cagrisi basarisiz.")


# ==================== 6. LLM CIKTISINI PARSE ETME ====================
def parse_json_candidate(text: str) -> tuple[int | None, float | None, bool]:
    parsed = json.loads(text)
    label = parsed.get("label")
    confidence = parsed.get("confidence")
    if isinstance(label, str):
        label = int(label.strip())
    if label not in {0, 1}:
        return None, None, False
    try:
        confidence_value = float(confidence)
    except Exception:
        confidence_value = None
    if confidence_value is not None:
        confidence_value = max(0.0, min(1.0, confidence_value))
    return int(label), confidence_value, True


def parse_llm_output(text: str) -> tuple[int | None, float | None, bool]:
    cleaned = (text or "").strip()
    cleaned = re.sub(r"<think>.*?</think>", "", cleaned, flags=re.DOTALL | re.IGNORECASE).strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?", "", cleaned).strip()
        cleaned = re.sub(r"```$", "", cleaned).strip()
    json_candidates = re.findall(r"\{[^{}]*\}", cleaned, flags=re.DOTALL)
    for candidate in reversed(json_candidates):
        try:
            return parse_json_candidate(candidate)
        except Exception:
            continue
    try:
        return parse_json_candidate(cleaned)
    except Exception:
        label_match = re.search(r"(?:^|\b)(?:label|etiket)\s*[:=]\s*([01])(?:\b|$)", cleaned, flags=re.IGNORECASE | re.MULTILINE)
        if label_match:
            return int(label_match.group(1)), None, False
    return None, None, False


# ==================== 7. PROGRESS VE ESKI SONUC OKUMA ====================
def output_sheet_name(model_key: str, scope: str, requested: str | None = None) -> str:
    if requested:
        return requested
    if model_key == "70b":
        return "LLM_Groq_Llama70B"
    if model_key == "qwen":
        return "LLM_Groq_Qwen"
    if model_key == "gptoss120b":
        return "LLM_Groq_GPTOSS120B"
    if model_key == "mistral":
        return "LLM_Mistral_Medium35"
    if model_key == "gemma":
        return "LLM_Gemini_Gemma4_31B"
    return "LLM_Groq_Llama8B"


def safe_model_name(model: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", model)


def progress_roots() -> list[Path]:
    return [PROGRESS_DIR]


def progress_path(sheet_name: str, model: str, scope: str, repeat: int) -> Path:
    return PROGRESS_DIR / f"{sheet_name}_{scope}_repeat{repeat}_{safe_model_name(model)}.jsonl"


def progress_read_candidates(sheet_name: str, model: str, scope: str, repeat: int) -> list[Path]:
    safe = safe_model_name(model)
    candidates: list[Path] = []
    scope_names = [scope]
    legacy_test_scope = LEGACY_TEST_SPLIT
    if normalize_split_name(scope) == TEST_SPLIT and legacy_test_scope not in scope_names:
        scope_names.append(legacy_test_scope)
    for root in progress_roots():
        for scope_name in scope_names:
            candidates.append(root / f"{sheet_name}_{scope_name}_repeat{repeat}_{safe}.jsonl")
            if repeat == 1:
                candidates.append(root / f"{sheet_name}_{scope_name}_{safe}.jsonl")
    return candidates


def read_progress_file(path: Path, repeat: int) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    records = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        record = json.loads(line)
        record["repeat"] = int(record.get("repeat") or repeat)
        record["scope"] = normalize_split_name(record.get("scope"))
        if record.get("raw_output"):
            label, confidence, parsed_ok = parse_llm_output(cell_text(record.get("raw_output")))
            record["llm_label"] = label
            record["confidence"] = confidence
            record["parsed_ok"] = parsed_ok
            manual = normalize_label(record.get("manual_label"))
            record["correct"] = int(label == manual) if label is not None and manual is not None else 0
        records.append(record)
    return records


def read_progress_records(sheet_name: str, model: str, scope: str) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    seen_paths: set[Path] = set()
    for repeat in (1, 2, 3):
        for path in progress_read_candidates(sheet_name, model, scope, repeat):
            if path in seen_paths:
                continue
            seen_paths.add(path)
            records.extend(read_progress_file(path, repeat))
    return records


def find_prediction_sections(sheet: Any) -> list[tuple[int, int]]:
    sections = []
    for row in range(1, sheet.max_row + 1):
        first = cell_text(sheet.cell(row, 1).value)
        if first.startswith("Predictions - Repeat") or first.startswith("Tahminler"):
            match = re.search(r"(?:Repeat|Tekrar)\s+([123])", first)
            repeat = int(match.group(1)) if match else 1
            sections.append((row, repeat))
    return sections


def read_existing_sheet_records(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    if not workbook_path.exists():
        return []
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        return []
    sheet = workbook[sheet_name]
    records: list[dict[str, Any]] = []
    sections = find_prediction_sections(sheet)
    for section_row, repeat in sections:
        header_row = section_row + 1
        headers = [cell_text(sheet.cell(header_row, col).value) for col in range(1, sheet.max_column + 1)]
        review_id_header = "review_id" if "review_id" in headers else "Review ID"
        sample_id_header = "sample_id" if "sample_id" in headers else "orneklem_id"
        meb_header = "Meb_Book_Name" if "Meb_Book_Name" in headers else "MEB_work" if "MEB_work" in headers else "MEB eser"
        if review_id_header not in headers or "llm_label" not in headers:
            continue
        for row in range(header_row + 1, sheet.max_row + 1):
            first = cell_text(sheet.cell(row, 1).value)
            if not first or first.startswith(("Ozet", "Summary", "Tahminler", "Predictions", "Cokluk", "Majority")):
                break
            item = {header: sheet.cell(row, col).value for col, header in enumerate(headers, start=1) if header}
            if not cell_text(item.get(review_id_header)):
                continue
            record = {
                "scope": TEST_SPLIT,
                "prompt_type": cell_text(item.get("prompt_type")) or "zero_shot",
                "fold": None,
                "repeat": int(item.get("repeat") or repeat),
                "orneklem_id": cell_text(item.get(sample_id_header)),
                "Review ID": cell_text(item.get(review_id_header)),
                "MEB eser": cell_text(item.get(meb_header)),
                "Goodreads eser": cell_text(item.get("Goodreads eser")),
                "yorum_raw": cell_text(item.get("yorum_raw")),
                "manual_label": normalize_label(item.get("human_label")),
                "llm_label": int(item["llm_label"]) if cell_text(item.get("llm_label")) in {"0", "1"} else None,
                "confidence": item.get("confidence"),
                "parsed_ok": bool(item.get("valid_output")),
                "correct": int(item.get("correct") or 0),
                "raw_output": "",
                "model": cell_text(item.get("model_name")),
            }
            records.append(record)
    workbook.close()
    return records


def enrich_records(records: list[dict[str, Any]], dataset_rows: list[dict[str, Any]]) -> None:
    rows_by_review_id = {cell_text(row.get("Review ID")): row for row in dataset_rows}
    for record in records:
        review_id = cell_text(record.get("Review ID"))
        source_row = rows_by_review_id.get(review_id)
        if not source_row:
            continue
        record["orneklem_id"] = cell_text(record.get("orneklem_id")) or cell_text(source_row.get("orneklem_id"))
        record["MEB eser"] = cell_text(record.get("MEB eser")) or cell_text(source_row.get("MEB eser"))
        record["Goodreads eser"] = cell_text(record.get("Goodreads eser")) or cell_text(source_row.get("Goodreads eser"))
        record["yorum_raw"] = cell_text(record.get("yorum_raw")) or cell_text(source_row.get("yorum_raw"))
        if "manual_label" not in record or record["manual_label"] in ("", None):
            record["manual_label"] = normalize_label(source_row.get("manuel_informative"))


def dedupe_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_key: dict[tuple[int, str], dict[str, Any]] = {}
    for record in records:
        if record.get("llm_label") not in {0, 1}:
            continue
        review_id = cell_text(record.get("Review ID"))
        if not review_id:
            continue
        repeat = int(record.get("repeat") or 1)
        by_key[(repeat, review_id)] = record
    return list(by_key.values())


def record_key(record: dict[str, Any]) -> tuple[int, str]:
    return (int(record.get("repeat") or 1), cell_text(record.get("Review ID")))


# ==================== 8. METRIK VE STABILITY ====================
def metrics(records: list[dict[str, Any]]) -> dict[str, Any]:
    valid = [record for record in records if record.get("llm_label") in {0, 1}]
    y_true = [int(record["manual_label"]) for record in valid]
    y_pred = [int(record["llm_label"]) for record in valid]
    tn = sum(1 for true, pred in zip(y_true, y_pred) if true == 0 and pred == 0)
    fp = sum(1 for true, pred in zip(y_true, y_pred) if true == 0 and pred == 1)
    fn = sum(1 for true, pred in zip(y_true, y_pred) if true == 1 and pred == 0)
    tp = sum(1 for true, pred in zip(y_true, y_pred) if true == 1 and pred == 1)
    accuracy = (tp + tn) / len(valid) if valid else 0
    precision = tp / (tp + fp) if (tp + fp) else 0
    recall = tp / (tp + fn) if (tp + fn) else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) else 0
    return {
        "n": len(records),
        "valid_n": len(valid),
        "accuracy": accuracy,
        "f1": f1,
        "precision": precision,
        "recall": recall,
        "correct": sum(1 for a, b in zip(y_true, y_pred) if a == b),
        "fp": int(fp),
        "fn": int(fn),
        "tn": int(tn),
        "tp": int(tp),
    }


def repeat_average_summary(records: list[dict[str, Any]], dataset_rows: list[dict[str, Any]]) -> dict[str, Any]:
    repeat_metrics = []
    for repeat in (1, 2, 3):
        repeat_records = sorted_records(records, dataset_rows, repeat)
        repeat_metrics.append(metrics(repeat_records))

    def avg(key: str) -> float:
        return sum(float(item[key]) for item in repeat_metrics) / len(repeat_metrics)

    evaluated_n = min(item["valid_n"] for item in repeat_metrics) if repeat_metrics else 0
    return {
        "evaluated_n": evaluated_n,
        "summary_rule": "average_of_3_repeats",
        "avg_accuracy": avg("accuracy"),
        "avg_f1": avg("f1"),
        "avg_precision": avg("precision"),
        "avg_recall": avg("recall"),
        "avg_correct": avg("correct"),
        "avg_fp": avg("fp"),
        "avg_fn": avg("fn"),
    }


def majority_records(records: list[dict[str, Any]], dataset_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_review: dict[str, dict[int, dict[str, Any]]] = {}
    for record in records:
        if record.get("llm_label") not in {0, 1}:
            continue
        review_id = cell_text(record.get("Review ID"))
        by_review.setdefault(review_id, {})[int(record.get("repeat") or 1)] = record

    source_by_id = {cell_text(row.get("Review ID")): row for row in dataset_rows}
    output = []
    for review_id, repeat_map in by_review.items():
        if not all(repeat in repeat_map for repeat in (1, 2, 3)):
            continue
        labels = [int(repeat_map[repeat]["llm_label"]) for repeat in (1, 2, 3)]
        counts = Counter(labels)
        majority_label, majority_count = counts.most_common(1)[0]
        source = source_by_id.get(review_id, {})
        manual = normalize_label(source.get("manuel_informative", repeat_map[1].get("manual_label")))
        stability_score = majority_count / 3
        output.append({
            "scope": TEST_SPLIT,
            "prompt_type": "zero_shot",
            "repeat": "majority",
            "orneklem_id": cell_text(source.get("orneklem_id")) or cell_text(repeat_map[1].get("orneklem_id")),
            "Review ID": review_id,
            "MEB eser": cell_text(source.get("MEB eser")) or cell_text(repeat_map[1].get("MEB eser")),
            "Goodreads eser": cell_text(source.get("Goodreads eser")) or cell_text(repeat_map[1].get("Goodreads eser")),
            "yorum_raw": cell_text(source.get("yorum_raw")) or cell_text(repeat_map[1].get("yorum_raw")),
            "manual_label": manual,
            "llm_label": int(majority_label),
            "confidence": None,
            "parsed_ok": True,
            "correct": int(int(majority_label) == manual),
            "model": cell_text(repeat_map[1].get("model")),
            "repeat_1_label": labels[0],
            "repeat_2_label": labels[1],
            "repeat_3_label": labels[2],
            "stability_type": "fully_stable_3_of_3" if majority_count == 3 else "partially_stable_2_of_3",
            "stability_score": stability_score,
        })
    return output


# ==================== 9. LLM CALISTIRMA ====================
def provider_config(args: argparse.Namespace) -> tuple[str, str, str, str]:
    model = args.model or DEFAULT_MODELS[args.model_key]
    provider = "Groq"
    api_url = API_URL
    api_key_name = "GROQ_API_KEY"
    if args.model_key == "mistral" or model.startswith("mistral-"):
        provider = "Mistral"
        api_url = MISTRAL_API_URL
        api_key_name = "MISTRAL_API_KEY"
    if args.model_key == "gemma" or model.startswith("gemma-"):
        provider = "Gemini"
        api_key_name = "GEMINI_API_KEY"
    return model, provider, api_url, api_key_name


def run_llm(args: argparse.Namespace) -> tuple[list[dict[str, Any]], list[dict[str, Any]], str, str]:
    load_env_files()
    model, provider, api_url, api_key_name = provider_config(args)
    api_key = os.getenv(api_key_name)
    if not api_key and not args.dry_run:
        searched = ", ".join(str(path) for path in env_paths())
        raise RuntimeError(f"{api_key_name} bulunamadi. Aranan .env dosyalari: {searched}")

    sheet_name = output_sheet_name(args.model_key, args.scope, args.output_sheet)
    write_progress = progress_path(sheet_name, model, args.scope, args.repeat)
    PROGRESS_DIR.mkdir(exist_ok=True)
    if args.clear_progress and write_progress.exists():
        write_progress.unlink()

    dataset_rows = read_dataset(args.workbook)
    base_rows = test_rows(dataset_rows)
    existing = []
    existing.extend(read_existing_sheet_records(args.workbook, sheet_name))
    existing.extend(read_progress_records(sheet_name, model, args.scope))
    enrich_records(existing, dataset_rows)
    existing = dedupe_records(existing)

    seen_current_repeat = {
        record_key(record)
        for record in existing
        if int(record.get("repeat") or 1) == args.repeat and record.get("llm_label") in {0, 1}
    }
    items = [{"scope": args.scope, "prompt_type": "zero_shot", "fold": None, "row": row} for row in base_rows]
    remaining = [
        item for item in items
        if (args.repeat, cell_text(item["row"].get("Review ID"))) not in seen_current_repeat
    ]

    print(f"Model: {model}")
    print(f"API: {provider}")
    print(f"Scope: {args.scope}")
    print(f"Prompt: zero_shot")
    print(f"Tekrar: {args.repeat}")
    print(f"Toplam test: {len(items)}; bu tekrar mevcut: {len(items) - len(remaining)}; kalan: {len(remaining)}")
    print(f"Progress: {write_progress}")

    if args.dry_run:
        sample = items[0]
        print("\nDRY RUN PROMPT ORNEGI:\n")
        print(build_prompt(sample["row"])[:1600])
        return existing, dataset_rows, sheet_name, model

    new_count = 0
    with write_progress.open("a", encoding="utf-8") as handle:
        for index, item in enumerate(items, start=1):
            row = item["row"]
            key = (args.repeat, cell_text(row.get("Review ID")))
            if key in seen_current_repeat:
                continue
            if args.max_new_records is not None and new_count >= args.max_new_records:
                break
            prompt = build_prompt(row)
            if provider == "Gemini":
                raw = call_gemini(api_key, model, prompt, args.sleep)
            else:
                raw = call_groq(api_key, model, prompt, args.sleep, api_url)
            label, confidence, parsed_ok = parse_llm_output(raw)
            manual = normalize_label(row.get("manuel_informative"))
            record = {
                "scope": item["scope"],
                "prompt_type": item["prompt_type"],
                "fold": item.get("fold"),
                "repeat": args.repeat,
                "orneklem_id": cell_text(row.get("orneklem_id")),
                "Review ID": cell_text(row.get("Review ID")),
                "MEB eser": cell_text(row.get("MEB eser")),
                "Goodreads eser": cell_text(row.get("Goodreads eser")),
                "yorum_raw": cell_text(row.get("yorum_raw")),
                "manual_label": manual,
                "llm_label": label,
                "confidence": confidence,
                "parsed_ok": parsed_ok,
                "correct": int(label == manual) if label is not None else 0,
                "raw_output": raw,
                "model": model,
            }
            handle.write(json.dumps(record, ensure_ascii=False) + "\n")
            handle.flush()
            existing.append(record)
            seen_current_repeat.add(key)
            new_count += 1
            if new_count % 10 == 0:
                print(f"  tekrar={args.repeat}, yeni={new_count}, son_index={index}/{len(items)}")

    existing = dedupe_records(existing)
    current_repeat_n = sum(1 for record in existing if int(record.get("repeat") or 1) == args.repeat)
    if current_repeat_n != len(items):
        print(f"UYARI: Tekrar {args.repeat} icin beklenen {len(items)}, mevcut {current_repeat_n}.")
    return existing, dataset_rows, sheet_name, model


# ==================== 10. EXCEL SAYFASI YAZMA ====================
def sheet_title(sheet_name: str) -> str:
    titles = {
        "LLM_Groq_Qwen": "Groq Qwen Informative LLM Deneyi",
        "LLM_Groq_GPTOSS120B": "Groq GPT-OSS 120B Informative LLM Deneyi",
        "LLM_Mistral_Medium35": "Mistral Medium 3.5 Informative LLM Deneyi",
        "LLM_Mistral_Large": "Mistral Large Informative LLM Deneyi",
        "LLM_Gemini_Gemma4_31B": "Gemini API Gemma 4 31B Informative LLM Deneyi",
        "LLM_Groq_Llama70B": "Groq Llama 70B Informative LLM Deneyi",
        "LLM_Groq_Llama8B": "Groq Llama 8B Informative LLM Deneyi",
    }
    return titles.get(sheet_name, "Informative LLM Deneyi")


def model_notes(sheet_name: str) -> list[list[str]]:
    notes = []
    if sheet_name == "LLM_Groq_Qwen":
        notes.append(["Qwen note", "Reasoning was disabled to return valid JSON without <think> text."])
    if sheet_name == "LLM_Groq_GPTOSS120B":
        notes.append(["GPT-OSS note", "Output budget was kept higher because it is a reasoning model; final answer is expected as JSON only."])
    if sheet_name == "LLM_Mistral_Medium35":
        notes.append(["Mistral note", "Mistral Medium 3.5 was selected because it supports JSON mode."])
    if sheet_name == "LLM_Gemini_Gemma4_31B":
        notes.append(["Gemma note", "Gemma 4 31B was tested through Gemini API on the test set."])
    return notes


def sorted_records(records: list[dict[str, Any]], dataset_rows: list[dict[str, Any]], repeat: int) -> list[dict[str, Any]]:
    order = {cell_text(row.get("Review ID")): index for index, row in enumerate(test_rows(dataset_rows))}
    subset = [record for record in records if int(record.get("repeat") or 1) == repeat]
    return sorted(subset, key=lambda item: order.get(cell_text(item.get("Review ID")), 10_000))


def write_repeat_section(sheet: Any, title: str, records: list[dict[str, Any]]) -> None:
    sheet.append([])
    sheet.append([title])
    header = [
        "repeat", "sample_id", "review_id", "Meb_Book_Name",
        "human_label", "llm_label", "correct", "confidence", "valid_output",
        "model_name", "prompt_type", "yorum_raw",
    ]
    sheet.append(header)
    for record in records:
        sheet.append([
            record.get("repeat"),
            record.get("orneklem_id"),
            record.get("Review ID"),
            record.get("MEB eser"),
            record.get("manual_label"),
            record.get("llm_label"),
            record.get("correct"),
            record.get("confidence"),
            record.get("parsed_ok"),
            record.get("model"),
            record.get("prompt_type", "zero_shot"),
            record.get("yorum_raw"),
        ])


def write_majority_section(sheet: Any, records: list[dict[str, Any]]) -> None:
    sheet.append([])
    sheet.append(["Majority Vote Predictions"])
    header = [
        "sample_id", "review_id", "Meb_Book_Name", "human_label",
        "repeat_1_label", "repeat_2_label", "repeat_3_label",
        "final_llm_label", "correct", "stability_type", "stability_score",
        "model_name", "prompt_type", "yorum_raw",
    ]
    sheet.append(header)
    for record in records:
        sheet.append([
            record.get("orneklem_id"),
            record.get("Review ID"),
            record.get("MEB eser"),
            record.get("manual_label"),
            record.get("repeat_1_label"),
            record.get("repeat_2_label"),
            record.get("repeat_3_label"),
            record.get("llm_label"),
            record.get("correct"),
            record.get("stability_type"),
            record.get("stability_score"),
            record.get("model"),
            record.get("prompt_type", "zero_shot"),
            record.get("yorum_raw"),
        ])


def write_excel(workbook_path: Path, sheet_name: str, records: list[dict[str, Any]], dataset_rows: list[dict[str, Any]]) -> None:
    workbook = load_workbook(workbook_path)
    for obsolete_sheet in [f"LLM_70b_{LEGACY_TEST_SPLIT}_ZS_V3", f"LLM_8b_{LEGACY_TEST_SPLIT}_ZS_V3", "LLM_Groq_Llama"]:
        if obsolete_sheet in workbook.sheetnames:
            del workbook[obsolete_sheet]
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    sheet = workbook.create_sheet(sheet_name)

    all_records = dedupe_records(records)
    model_name = cell_text(all_records[0].get("model")) if all_records else ""

    sheet.append([sheet_title(sheet_name)])
    sheet.append(["Dataset", "Informative_ML_Dataset / split=test"])
    sheet.append(["Text Column", "yorum_raw"])
    sheet.append(["Evaluation", "The same 200 test reviews are classified 3 times with zero-shot prompting."])
    sheet.append(["Prompt Type", "zero_shot"])
    sheet.append(["Repeat Design", "Repeats 1/2/3 are stored separately; final summary uses the arithmetic average of repeat metrics."])
    sheet.append(["Reason Output", "none"])
    for note in model_notes(sheet_name):
        sheet.append(note)

    sheet.append([])
    sheet.append(["Repeat Summary"])
    sheet.append(["model_name", "prompt_type", "repeat", "total_n", "valid_n", "accuracy", "f1", "precision", "recall", "correct", "fp", "fn"])
    for repeat in (1, 2, 3):
        repeat_records = sorted_records(all_records, dataset_rows, repeat)
        m = metrics(repeat_records)
        sheet.append([
            model_name,
            "zero_shot",
            repeat,
            m["n"],
            m["valid_n"],
            round(m["accuracy"], 4),
            round(m["f1"], 4),
            round(m["precision"], 4),
            round(m["recall"], 4),
            f"{m['correct']}/{m['valid_n']}",
            m["fp"],
            m["fn"],
        ])

    average_summary = repeat_average_summary(all_records, dataset_rows)
    sheet.append([])
    sheet.append(["Repeat Average Summary"])
    sheet.append([
        "model_name", "prompt_type", "evaluated_n", "summary_rule",
        "avg_accuracy", "avg_f1", "avg_precision", "avg_recall", "avg_correct", "avg_fp", "avg_fn",
    ])
    sheet.append([
        model_name,
        "zero_shot",
        average_summary["evaluated_n"],
        average_summary["summary_rule"],
        round(average_summary["avg_accuracy"], 4),
        round(average_summary["avg_f1"], 4),
        round(average_summary["avg_precision"], 4),
        round(average_summary["avg_recall"], 4),
        f"{average_summary['avg_correct']:.2f}/{average_summary['evaluated_n']}",
        round(average_summary["avg_fp"], 2),
        round(average_summary["avg_fn"], 2),
    ])

    for repeat in (1, 2, 3):
        write_repeat_section(sheet, f"Predictions - Repeat {repeat}", sorted_records(all_records, dataset_rows, repeat))

    style_llm_sheet(sheet)
    update_llm_comparison_sheet(workbook)
    workbook.active = workbook.sheetnames.index(sheet_name)
    workbook.save(workbook_path)


# ==================== 11. LLM KARSILASTIRMA SAYFASI ====================
def row_values(sheet: Any, row: int) -> list[Any]:
    return [sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)]


def find_row(sheet: Any, label: str) -> int | None:
    for row in range(1, sheet.max_row + 1):
        if cell_text(sheet.cell(row, 1).value) == label:
            return row
    return None


def extract_llm_summary(sheet: Any) -> dict[str, Any] | None:
    average_title_row = find_row(sheet, "Repeat Average Summary")
    majority_title_row = find_row(sheet, "Majority Vote and Stability Summary") or find_row(sheet, "Ozet - Cokluk Oyu ve Stability")
    repeat_title_row = find_row(sheet, "Repeat Summary") or find_row(sheet, "Ozet - Tekrarlar")
    if not average_title_row and not majority_title_row and not repeat_title_row:
        legacy_header = find_row(sheet, "model_name")
        if legacy_header:
            values = row_values(sheet, legacy_header + 1)
            if len(values) >= 11 and values[0]:
                return {
                    "model_name": values[0],
                    "prompt_type": values[1],
                    "evaluated_n": values[2],
                    "summary_rule": "single_run",
                    "avg_accuracy": values[4],
                    "avg_f1": values[5],
                    "avg_precision": values[6],
                    "avg_recall": values[7],
                    "avg_correct": values[8],
                    "avg_fp": values[9],
                    "avg_fn": values[10],
                }
        return None

    result: dict[str, Any] = {}
    if average_title_row:
        headers = [cell_text(value) for value in row_values(sheet, average_title_row + 1)]
        values = row_values(sheet, average_title_row + 2)
        result.update({header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header})
    elif majority_title_row:
        headers = [cell_text(value) for value in row_values(sheet, majority_title_row + 1)]
        values = row_values(sheet, majority_title_row + 2)
        legacy = {header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header}
        result.update({
            "model_name": legacy.get("model_name"),
            "prompt_type": legacy.get("prompt_type"),
            "evaluated_n": legacy.get("evaluated_n"),
            "summary_rule": legacy.get("final_decision_rule"),
            "avg_accuracy": legacy.get("accuracy"),
            "avg_f1": legacy.get("f1"),
            "avg_precision": legacy.get("precision"),
            "avg_recall": legacy.get("recall"),
            "avg_correct": legacy.get("correct"),
            "avg_fp": legacy.get("fp"),
            "avg_fn": legacy.get("fn"),
        })

    if repeat_title_row:
        headers = [cell_text(value) for value in row_values(sheet, repeat_title_row + 1)]
        for row in range(repeat_title_row + 2, repeat_title_row + 5):
            values = row_values(sheet, row)
            item = {header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header}
            repeat = cell_text(item.get("repeat"))
            if repeat in {"1", "2", "3"}:
                result[f"repeat_{repeat}_accuracy"] = item.get("accuracy") if "accuracy" in item else item.get("Accuracy")
                result[f"repeat_{repeat}_f1"] = item.get("f1") if "f1" in item else item.get("F1")
    return result if result.get("model_name") else None


def update_llm_comparison_sheet(workbook: Any) -> None:
    rows = []
    for sheet_name, provider, method_name in LLM_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        summary = extract_llm_summary(workbook[sheet_name])
        if not summary:
            continue
        summary["Yontem"] = method_name
        summary["Saglayici"] = provider
        rows.append(summary)

    def evaluated_n_value(item: dict[str, Any]) -> int:
        value = item.get("evaluated_n") if item.get("evaluated_n") is not None else item.get("Majority_N")
        try:
            return int(value)
        except Exception:
            return 0

    rows.sort(
        key=lambda item: (
            evaluated_n_value(item) >= 200,
            float(item.get("avg_f1") or item.get("repeat_1_f1") or 0),
            float(item.get("avg_accuracy") or item.get("repeat_1_accuracy") or 0),
        ),
        reverse=True,
    )

    def top_ml_rows() -> list[dict[str, Any]]:
        if "Informative_Model_Karsilastirma" not in workbook.sheetnames:
            return []
        source = workbook["Informative_Model_Karsilastirma"]
        section_row = (
            find_row(source, "GridSearchCV Results")
            or find_row(source, "GridSearchCV Top 5 Results")
            or find_row(source, "Grid Search Top 5 Sonuclari")
            or find_row(source, "Fixed Method Results")
            or find_row(source, "Sabit Yontem Sonuclari")
        )
        header_row = section_row + 1 if section_row else None
        if not header_row:
            return []
        headers = [cell_text(value) for value in row_values(source, header_row)]
        ml_rows = []
        for row_index in range(header_row + 1, source.max_row + 1):
            first = cell_text(source.cell(row_index, 1).value)
            if not first or first.startswith("Grid"):
                if ml_rows:
                    break
                continue
            item = {header: source.cell(row_index, col).value for col, header in enumerate(headers, start=1) if header}
            if item.get("Classification Method"):
                item["method_label"] = f"{item.get('Classification Method')} + {item.get('Text Representation')}"
                item["representation_label"] = item.get("Text Representation")
                item["model_setting"] = item.get("Classification Parameters")
                item["representation_setting"] = item.get("Representation Parameters")
                ml_rows.append(item)
            elif item.get("Yontem"):
                item["method_label"] = item.get("Yontem")
                item["representation_label"] = item.get("Temsil ayari")
                item["model_setting"] = item.get("Model ayari")
                item["representation_setting"] = item.get("Temsil ayari")
                ml_rows.append(item)
        ml_rows.sort(
            key=lambda item: (
                float(item.get("Final F1") or item.get("Final F1-score") or 0),
                float(item.get("Final Accuracy") or 0),
                float(item.get("10-Fold F1") or 0),
            ),
            reverse=True,
        )
        return ml_rows[:5]

    if "LLM_Karsilastirma" in workbook.sheetnames:
        del workbook["LLM_Karsilastirma"]
    if "ML_LLM_Karsilastirma" in workbook.sheetnames:
        del workbook["ML_LLM_Karsilastirma"]
    sheet = workbook.create_sheet("ML_LLM_Karsilastirma")
    sheet.append(["ML + LLM Comparison - Test"])
    sheet.append(["Scope", "LLM results use the arithmetic average of 3 zero-shot repeats; ML results use test performance."])
    sheet.append(["ML ranking", "Top 5 ML methods are selected by Final F1; 10-fold metrics are shown for reference."])
    sheet.append([])
    sheet.append(["LLM Results - Repeat Average"])
    headers = [
        "rank", "method", "provider", "model_name", "prompt_type",
        "evaluated_n", "status", "summary_rule", "avg_accuracy", "avg_f1",
        "avg_precision", "avg_recall", "avg_correct", "avg_fp", "avg_fn",
    ]
    sheet.append(headers)
    for index, row in enumerate(rows, start=1):
        def pick(primary: str, fallback: str) -> Any:
            value = row.get(primary)
            return value if value is not None and value != "" else row.get(fallback)

        sheet.append([
            index,
            row.get("Yontem"),
            row.get("Saglayici"),
            row.get("model_name"),
            row.get("prompt_type"),
            pick("evaluated_n", "Majority_N"),
            "complete" if evaluated_n_value(row) >= 200 else "interim_incomplete",
            row.get("summary_rule") or "average_of_3_repeats",
            pick("avg_accuracy", "accuracy"),
            pick("avg_f1", "f1"),
            pick("avg_precision", "precision"),
            pick("avg_recall", "recall"),
            pick("avg_correct", "correct"),
            pick("avg_fp", "fp"),
            pick("avg_fn", "fn"),
        ])
    ml_rows = top_ml_rows()
    if ml_rows:
        sheet.append([])
        sheet.append(["Top 5 ML Methods - Test"])
        sheet.append([
            "rank", "classification_method", "text_representation", "representation_setting", "model_setting",
            "10-fold_accuracy", "10-fold_f1", "final_accuracy", "final_f1",
            "final_precision", "final_recall", "correct", "fp", "fn",
        ])
        for index, row in enumerate(ml_rows, start=1):
            sheet.append([
                index,
                row.get("Classification Method") or row.get("method_label"),
                row.get("Text Representation") or row.get("representation_label"),
                row.get("representation_setting"),
                row.get("model_setting"),
                row.get("10-Fold Accuracy"),
                row.get("10-Fold F1"),
                row.get("Final Accuracy"),
                row.get("Final F1") or row.get("Final F1-score"),
                row.get("Final Precision"),
                row.get("Final Recall"),
                row.get("Final Correct") or row.get("Dogru"),
                row.get("FP"),
                row.get("FN"),
            ])
    style_comparison_sheet(sheet)


# ==================== 12. EXCEL GORUNUMU ====================
def style_llm_sheet(sheet: Any) -> None:
    title_fill = PatternFill("solid", fgColor="93C47D")
    section_fill = PatternFill("solid", fgColor="B6D7A8")
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    thin = Side(style="thin", color="DDDDDD")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = Border(bottom=thin)
    for row_index in range(1, sheet.max_row + 1):
        first = cell_text(sheet.cell(row_index, 1).value)
        if row_index == 1:
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row_index, col).font = Font(bold=True, size=13)
                sheet.cell(row_index, col).fill = title_fill
        if first in {
            "Repeat Summary",
            "Repeat Average Summary",
            "Predictions - Repeat 1",
            "Predictions - Repeat 2",
            "Predictions - Repeat 3",
            "Ozet - Tekrarlar",
            "Tahminler - Tekrar 1",
            "Tahminler - Tekrar 2",
            "Tahminler - Tekrar 3",
        }:
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row_index, col).font = Font(bold=True)
                sheet.cell(row_index, col).fill = section_fill
        if first in {"model_name", "repeat", "sample_id", "orneklem_id", "rank"}:
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row_index, col).font = Font(bold=True)
                sheet.cell(row_index, col).fill = header_fill

    widths = {
        "A": 13, "B": 14, "C": 34, "D": 26, "E": 12, "F": 10, "G": 10,
        "H": 12, "I": 12, "J": 28, "K": 14, "L": 80, "M": 14, "N": 80,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = "A14"


def style_comparison_sheet(sheet: Any) -> None:
    title_fill = PatternFill("solid", fgColor="93C47D")
    section_fill = PatternFill("solid", fgColor="B6D7A8")
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    thin = Side(style="thin", color="DDDDDD")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = Border(bottom=thin)
    for row_index in range(1, sheet.max_row + 1):
        first = cell_text(sheet.cell(row_index, 1).value)
        if row_index == 1:
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row_index, col).font = Font(bold=True, size=13)
                sheet.cell(row_index, col).fill = title_fill
        if first in {"LLM Results - Repeat Average", "Top 5 ML Methods - Test"}:
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row_index, col).font = Font(bold=True)
                sheet.cell(row_index, col).fill = section_fill
        if first == "rank":
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row_index, col).font = Font(bold=True)
                sheet.cell(row_index, col).fill = header_fill
    widths = {
        "A": 7, "B": 22, "C": 14, "D": 28, "E": 13, "F": 12, "G": 24,
        "H": 11, "I": 10, "J": 11, "K": 10, "L": 12, "M": 8, "N": 8,
        "O": 8, "P": 34, "Q": 34,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = "A7"


# ==================== 13. CALISTIRMA AKISI ====================
def main() -> None:
    args = parse_args()
    if args.update_comparison_only:
        workbook = load_workbook(args.workbook)
        update_llm_comparison_sheet(workbook)
        workbook.save(args.workbook)
        print("ML_LLM_Karsilastirma updated.")
        return
    records, dataset_rows, sheet_name, _model = run_llm(args)
    if not args.dry_run:
        write_excel(args.workbook, sheet_name, records, dataset_rows)
        print(f"Excel sheet written: {sheet_name}")
        print("Repeat average summary updated.")


if __name__ == "__main__":
    main()
