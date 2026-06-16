from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sklearn.model_selection import StratifiedKFold


# ==================== 1. DOSYA VE KOLON AYARLARI ====================
BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "MEB_100_Goodreads_ML_LLM.xlsx"
DATASET_SHEET = "Informative_ML_Dataset"

ID_COLUMN = "Review ID"
LABEL_COLUMN = "manuel_informative"
SPLIT_COLUMN = "split"
FOLD_COLUMN = "cv_fold(10)"
RANDOM_STATE = 42
TRAIN_SPLIT = "train"
TEST_SPLIT = "test"
LEGACY_TRAIN_SPLIT = "train" + "_cv"
LEGACY_TEST_SPLIT = "final" + "_test"
TRAIN_SPLIT_ALIASES = {TRAIN_SPLIT, LEGACY_TRAIN_SPLIT}
TEST_SPLIT_ALIASES = {TEST_SPLIT, LEGACY_TEST_SPLIT}


# ==================== 2. YARDIMCI FONKSIYONLAR ====================
def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_label(value: Any) -> str:
    text = cell_text(value)
    if text in {"0", "0.0"}:
        return "0"
    if text in {"1", "1.0"}:
        return "1"
    raise ValueError(f"Beklenmeyen informative etiketi: {value!r}")


def header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }


def ensure_fold_column(sheet) -> dict[str, int]:
    columns = header_map(sheet)
    if FOLD_COLUMN not in columns:
        insert_at = columns[SPLIT_COLUMN] + 1
        sheet.insert_cols(insert_at)
        sheet.cell(1, insert_at).value = FOLD_COLUMN
    return header_map(sheet)


# ==================== 3. 10-FOLD ATAMA ====================
def assign_10fold(sheet) -> dict[int, Counter]:
    columns = ensure_fold_column(sheet)
    required = [ID_COLUMN, LABEL_COLUMN, SPLIT_COLUMN, FOLD_COLUMN]
    missing = [column for column in required if column not in columns]
    if missing:
        raise ValueError(f"Eksik kolonlar: {missing}")

    train_rows: list[int] = []
    labels: list[str] = []
    for row_no in range(2, sheet.max_row + 1):
        split_value = cell_text(sheet.cell(row_no, columns[SPLIT_COLUMN]).value)
        label = normalize_label(sheet.cell(row_no, columns[LABEL_COLUMN]).value)
        if split_value in TRAIN_SPLIT_ALIASES:
            sheet.cell(row_no, columns[SPLIT_COLUMN]).value = TRAIN_SPLIT
            train_rows.append(row_no)
            labels.append(label)
        elif split_value in TEST_SPLIT_ALIASES:
            sheet.cell(row_no, columns[SPLIT_COLUMN]).value = TEST_SPLIT
            sheet.cell(row_no, columns[FOLD_COLUMN]).value = None

    splitter = StratifiedKFold(n_splits=10, shuffle=True, random_state=RANDOM_STATE)
    fold_summary: dict[int, Counter] = {}
    for fold_no, (_, valid_index) in enumerate(splitter.split(train_rows, labels), start=1):
        fold_summary[fold_no] = Counter(labels[index] for index in valid_index)
        for index in valid_index:
            sheet.cell(train_rows[index], columns[FOLD_COLUMN]).value = fold_no
    return dict(sorted(fold_summary.items()))


# ==================== 4. ANA AKIS ====================
def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH)
    if DATASET_SHEET not in workbook.sheetnames:
        raise ValueError(f"Sayfa bulunamadi: {DATASET_SHEET}")

    summary = assign_10fold(workbook[DATASET_SHEET])
    workbook.save(WORKBOOK_PATH)

    print("10-fold atama tamamlandi.")
    for fold, counts in summary.items():
        print(f"fold {fold}: 0={counts['0']}, 1={counts['1']}, total={sum(counts.values())}")
    print(f"Workbook updated: {WORKBOOK_PATH}")


if __name__ == "__main__":
    main()
