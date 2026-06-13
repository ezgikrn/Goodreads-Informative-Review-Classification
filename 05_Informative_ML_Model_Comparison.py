from __future__ import annotations

import argparse
import json
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from statistics import mean, stdev
from typing import Any

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sklearn.base import clone
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
from sklearn.linear_model import LogisticRegression, RidgeClassifier
from sklearn.metrics import accuracy_score, confusion_matrix, f1_score, make_scorer, precision_score, recall_score
from sklearn.model_selection import GridSearchCV, PredefinedSplit
from sklearn.naive_bayes import GaussianNB, MultinomialNB
from sklearn.neighbors import KNeighborsClassifier
from sklearn.pipeline import FeatureUnion, Pipeline
from sklearn.preprocessing import FunctionTransformer, Normalizer, StandardScaler
from sklearn.svm import SVC


# ==================== 1. DOSYA VE KOLON AYARLARI ====================
BASE_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = BASE_DIR / "MEB_100_Goodreads_ML_LLM.xlsx"
DATASET_SHEET = "Informative_ML_Dataset"
CORPUS_SHEET = "Puan_Analizi"
RESULTS_SHEET = "Informative_Model_Karsilastirma"

ID_COLUMN = "Review ID"
LABEL_COLUMN = "manuel_informative"
SPLIT_COLUMN = "split"
FOLD_COLUMN = "cv_fold(10)"
TEXT_COLUMN = "yorum_temiz"

REQUIRED_COLUMNS = [
    ID_COLUMN,
    "orneklem_id",
    "yorum_raw",
    TEXT_COLUMN,
    LABEL_COLUMN,
    SPLIT_COLUMN,
    FOLD_COLUMN,
]

RANDOM_STATE = 42
TRAIN_SPLIT = "train"
TEST_SPLIT = "test"
LEGACY_TRAIN_SPLIT = "train" + "_cv"
LEGACY_TEST_SPLIT = "final" + "_test"
TRAIN_SPLIT_ALIASES = {TRAIN_SPLIT, LEGACY_TRAIN_SPLIT}
TEST_SPLIT_ALIASES = {TEST_SPLIT, LEGACY_TEST_SPLIT}


# ==================== 2. ORTAK GRID PARAMETRE ALANLARI ====================
LOGISTIC_C_VALUES = [0.1, 0.25, 0.5, 0.75, 1.0, 2.0, 4.0]
SVM_C_VALUES = [0.1, 0.25, 0.5, 0.75, 1.0, 2.0]
SVM_KERNEL_VALUES = ["linear", "rbf", "poly"]
SVM_GAMMA_VALUES = ["scale", "auto"]
SVM_DEGREE_VALUES = [2, 3]
RIDGE_ALPHA_VALUES = [0.1, 0.5, 1.0, 2.0, 5.0, 10.0]
NB_ALPHA_VALUES = [0.01, 0.05, 0.1, 0.5, 1.0]
CLASS_WEIGHT_VALUES = ["balanced", None]
K_VALUES = [3, 5, 7, 9, 11, 15, 21, 31]
KNN_WEIGHTS = ["uniform", "distance"]
BERT_KNN_TRANSFORMS = ["none", "l2_normalize", "standard_scale"]


# ==================== 3. TEMSIL MODEL AYARLARI ====================
W2V_VECTOR_SIZE = 100
W2V_WINDOW = 5
W2V_MIN_COUNT = 2
W2V_EPOCHS = 30

FASTTEXT_VECTOR_SIZE = 100
FASTTEXT_WINDOW = 5
FASTTEXT_MIN_COUNT = 2
FASTTEXT_EPOCHS = 30

BERT_MODEL_NAME = "dbmdz/bert-base-turkish-cased"
BERT_MODEL_LABEL = "BERT dbmdz/bert-base-turkish-cased mean pooling"
BERT_CACHE_DIR = BASE_DIR / "_bert_cache"
BERT_EMBEDDING_PATH = BERT_CACHE_DIR / "dbmdz_bert_base_turkish_cased_mean_1000.npy"
BERT_IDS_PATH = BERT_CACHE_DIR / "dbmdz_bert_base_turkish_cased_mean_1000_ids.json"
BERT_MAX_LENGTH = 256
BERT_BATCH_SIZE = 8


# ==================== 4. YONTEM LISTESI ====================
@dataclass(frozen=True)
class MethodConfig:
    name: str
    vectorizer: str
    model: str
    params: dict[str, Any]


BASELINE_METHODS = [
    # Logistic Regression + 6 text representations
    MethodConfig("Logistic Regression + TF-IDF", "tfidf", "logistic_regression", {
        "ngram_range": (1, 1), "min_df": 2, "max_df": 0.90, "C": 1.0, "class_weight": "balanced",
    }),
    MethodConfig("Logistic Regression + TF-IDF + Character N-gram", "tfidf_word_char", "logistic_regression", {
        "word_ngram_range": (1, 1), "char_ngram_range": (3, 6), "word_min_df": 2,
        "char_min_df": 2, "max_df": 0.90, "C": 2.0, "class_weight": "balanced",
    }),
    MethodConfig("Logistic Regression + BoW", "count", "logistic_regression", {
        "ngram_range": (1, 2), "min_df": 1, "max_df": 0.90, "C": 0.75, "class_weight": None,
    }),
    MethodConfig("Logistic Regression + Word2Vec", "word2vec", "logistic_regression", {
        "pooling": "mean", "C": 1.0, "class_weight": "balanced",
    }),
    MethodConfig("Logistic Regression + FastText", "fasttext", "logistic_regression", {
        "pooling": "mean", "C": 0.25, "class_weight": "balanced",
    }),
    MethodConfig("Logistic Regression + BERT", "bert", "logistic_regression", {
        "pooling": "mean", "max_length": BERT_MAX_LENGTH, "C": 0.25, "class_weight": None,
    }),

    # SVM + 6 text representations
    MethodConfig("SVM + TF-IDF", "tfidf", "linear_svm", {
        "ngram_range": (1, 1), "min_df": 2, "max_df": 0.90, "C": 1.0, "class_weight": "balanced",
    }),
    MethodConfig("SVM + TF-IDF + Character N-gram", "tfidf_word_char", "linear_svm", {
        "word_ngram_range": (1, 1), "char_ngram_range": (3, 6), "word_min_df": 2,
        "char_min_df": 2, "max_df": 0.90, "C": 0.25, "class_weight": "balanced",
    }),
    MethodConfig("SVM + BoW", "count", "linear_svm", {
        "ngram_range": (1, 2), "min_df": 1, "max_df": 0.90, "C": 1.0, "class_weight": "balanced",
    }),
    MethodConfig("SVM + Word2Vec", "word2vec", "linear_svm", {
        "pooling": "mean", "C": 0.1, "class_weight": "balanced",
    }),
    MethodConfig("SVM + FastText", "fasttext", "linear_svm", {
        "pooling": "mean", "C": 0.1, "class_weight": "balanced",
    }),
    MethodConfig("SVM + BERT", "bert", "linear_svm", {
        "pooling": "mean", "max_length": BERT_MAX_LENGTH, "C": 0.25, "class_weight": "balanced",
    }),

    # Ridge Classifier + 6 text representations
    MethodConfig("Ridge Classifier + TF-IDF", "tfidf", "ridge_classifier", {
        "ngram_range": (1, 1), "min_df": 2, "max_df": 0.90, "alpha": 5.0, "class_weight": "balanced",
    }),
    MethodConfig("Ridge Classifier + TF-IDF + Character N-gram", "tfidf_word_char", "ridge_classifier", {
        "word_ngram_range": (1, 1), "char_ngram_range": (3, 6), "word_min_df": 2,
        "char_min_df": 2, "max_df": 0.90, "alpha": 2.0, "class_weight": "balanced",
    }),
    MethodConfig("Ridge Classifier + BoW", "count", "ridge_classifier", {
        "ngram_range": (1, 2), "min_df": 1, "max_df": 0.90, "alpha": 2.0, "class_weight": "balanced",
    }),
    MethodConfig("Ridge Classifier + Word2Vec", "word2vec", "ridge_classifier", {
        "pooling": "mean", "alpha": 5.0, "class_weight": "balanced",
    }),
    MethodConfig("Ridge Classifier + FastText", "fasttext", "ridge_classifier", {
        "pooling": "mean", "alpha": 10.0, "class_weight": "balanced",
    }),
    MethodConfig("Ridge Classifier + BERT", "bert", "ridge_classifier", {
        "pooling": "mean", "max_length": BERT_MAX_LENGTH, "alpha": 10.0, "class_weight": None,
    }),

    # Naive Bayes + 6 text representations
    MethodConfig("Naive Bayes + TF-IDF", "tfidf", "naive_bayes", {
        "ngram_range": (1, 1), "min_df": 2, "max_df": 0.90, "alpha": 0.5,
    }),
    MethodConfig("Naive Bayes + TF-IDF + Character N-gram", "tfidf_word_char", "naive_bayes", {
        "word_ngram_range": (1, 1), "char_ngram_range": (3, 6), "word_min_df": 2,
        "char_min_df": 2, "max_df": 0.90, "alpha": 0.5,
    }),
    MethodConfig("Naive Bayes + BoW", "count", "naive_bayes", {
        "ngram_range": (1, 2), "min_df": 1, "max_df": 0.90, "alpha": 0.5,
    }),
    MethodConfig("Naive Bayes + Word2Vec", "word2vec", "naive_bayes", {
        "pooling": "mean", "variant": "GaussianNB",
    }),
    MethodConfig("Naive Bayes + FastText", "fasttext", "naive_bayes", {
        "pooling": "mean", "variant": "GaussianNB",
    }),
    MethodConfig("Naive Bayes + BERT", "bert", "naive_bayes", {
        "pooling": "mean", "max_length": BERT_MAX_LENGTH, "variant": "GaussianNB",
    }),

    # KNN + 6 text representations
    MethodConfig("KNN + TF-IDF", "tfidf", "knn", {
        "ngram_range": (1, 1), "min_df": 2, "max_df": 0.90,
        "n_neighbors": 3, "weights": "uniform", "metric": "cosine",
    }),
    MethodConfig("KNN + TF-IDF + Character N-gram", "tfidf_word_char", "knn", {
        "word_ngram_range": (1, 1), "char_ngram_range": (3, 6), "word_min_df": 2,
        "char_min_df": 2, "max_df": 0.90, "n_neighbors": 3, "weights": "uniform", "metric": "cosine",
    }),
    MethodConfig("KNN + BoW", "count", "knn", {
        "ngram_range": (1, 2), "min_df": 1, "max_df": 0.90,
        "n_neighbors": 3, "weights": "uniform", "metric": "cosine",
    }),
    MethodConfig("KNN + Word2Vec", "word2vec", "knn", {
        "pooling": "mean", "transform": "l2_normalize", "n_neighbors": 7, "weights": "uniform", "metric": "cosine",
    }),
    MethodConfig("KNN + FastText", "fasttext", "knn", {
        "pooling": "mean", "transform": "l2_normalize", "n_neighbors": 7, "weights": "uniform", "metric": "cosine",
    }),
    MethodConfig("KNN + BERT", "bert", "knn", {
        "pooling": "mean", "max_length": BERT_MAX_LENGTH, "transform": "standard_scale",
        "n_neighbors": 7, "weights": "uniform", "metric": "cosine",
    }),
]


# ==================== 5. KOMUT SATIRI ====================
def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Informative ML V3 - 10-fold ve test akisi.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--sheet", default=DATASET_SHEET)
    parser.add_argument("--mode", choices=["validate", "baseline", "grid_top5", "grid_top5_only", "grid_all"], default="validate")
    parser.add_argument("--text-col", default=TEXT_COLUMN)
    parser.add_argument("--top-n", type=int, default=5)
    parser.add_argument("--grid-method", default=None)
    parser.add_argument("--no-excel", action="store_true")
    return parser.parse_args()


# ==================== 6. VERI OKUMA VE KONTROL ====================
def read_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Sheet not found: {sheet_name}")
    sheet = workbook[sheet_name]
    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")
    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({header: values[index] if index < len(values) else None for index, header in enumerate(headers)})
    workbook.close()
    return rows


def label(value: Any) -> str:
    normalized = str(value).strip()
    if normalized not in {"0", "1"}:
        raise ValueError(f"Invalid label: {value!r}")
    return normalized


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_split_name(value: Any) -> str:
    split = cell_text(value)
    if split in TRAIN_SPLIT_ALIASES:
        return TRAIN_SPLIT
    if split in TEST_SPLIT_ALIASES:
        return TEST_SPLIT
    return split


def is_train_split(value: Any) -> bool:
    return normalize_split_name(value) == TRAIN_SPLIT


def is_test_split(value: Any) -> bool:
    return normalize_split_name(value) == TEST_SPLIT


def validate_dataset(rows: list[dict[str, Any]], text_col: str) -> None:
    if not rows:
        raise ValueError("Dataset is empty.")
    split_counts = Counter(normalize_split_name(row.get(SPLIT_COLUMN)) for row in rows)
    if set(split_counts) != {TRAIN_SPLIT, TEST_SPLIT}:
        raise ValueError(f"Unexpected split values: {dict(split_counts)}")
    if "cv_fold(5)" in rows[0]:
        raise ValueError("cv_fold(5) should not be used in V3 dataset.")

    train_rows = [row for row in rows if is_train_split(row.get(SPLIT_COLUMN))]
    test_rows = [row for row in rows if is_test_split(row.get(SPLIT_COLUMN))]
    folds = sorted({cell_text(row.get(FOLD_COLUMN)) for row in train_rows}, key=int)
    if folds != [str(i) for i in range(1, 11)]:
        raise ValueError(f"Expected 10 folds 1..10, found: {folds}")
    final_with_fold = [row.get(ID_COLUMN) for row in test_rows if cell_text(row.get(FOLD_COLUMN))]
    if final_with_fold:
        raise ValueError(f"test rows must have empty {FOLD_COLUMN}; count={len(final_with_fold)}")
    empty_text = [row.get(ID_COLUMN) for row in rows if not cell_text(row.get(text_col))]
    if empty_text:
        raise ValueError(f"Empty text rows in {text_col}; count={len(empty_text)}")

    print("Dataset kontrolu OK")
    print(f"Toplam={len(rows)} train={len(train_rows)} test={len(test_rows)}")
    print("Train label dagilimi:", dict(Counter(label(row[LABEL_COLUMN]) for row in train_rows)))
    print("Test label dagilimi:", dict(Counter(label(row[LABEL_COLUMN]) for row in test_rows)))
    print("10-fold dagilimi:", {fold: sum(1 for row in train_rows if cell_text(row[FOLD_COLUMN]) == fold) for fold in folds})


# ==================== 7. ORTAK SPLIT VE METRIKLER ====================
def train_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in rows if is_train_split(row.get(SPLIT_COLUMN))]


def test_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in rows if is_test_split(row.get(SPLIT_COLUMN))]


def to_y(rows: list[dict[str, Any]]) -> list[str]:
    return [label(row.get(LABEL_COLUMN)) for row in rows]


def metrics(y_true: list[str], y_pred: list[str]) -> dict[str, Any]:
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=["0", "1"]).ravel()
    correct = sum(1 for actual, predicted in zip(y_true, y_pred) if actual == predicted)
    return {
        "accuracy": accuracy_score(y_true, y_pred),
        "f1_score": f1_score(y_true, y_pred, pos_label="1", zero_division=0),
        "precision": precision_score(y_true, y_pred, pos_label="1", zero_division=0),
        "recall": recall_score(y_true, y_pred, pos_label="1", zero_division=0),
        "correct": correct,
        "total": len(y_true),
        "tn": int(tn),
        "fp": int(fp),
        "fn": int(fn),
        "tp": int(tp),
    }


def mean_metrics(fold_rows: list[dict[str, Any]]) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for key in ["accuracy", "f1_score", "precision", "recall"]:
        values = [float(row[key]) for row in fold_rows]
        output[f"fold_{key}"] = mean(values)
        output[f"fold_{key}_std"] = stdev(values) if len(values) > 1 else 0.0
    output["fold_correct"] = sum(int(row["correct"]) for row in fold_rows)
    output["fold_total"] = sum(int(row["total"]) for row in fold_rows)
    return output


def predefined_10fold(train_rows: list[dict[str, Any]]) -> PredefinedSplit:
    return PredefinedSplit([int(cell_text(row.get(FOLD_COLUMN))) - 1 for row in train_rows])


# ==================== 8. TEMSIL HAZIRLIGI ====================
def tokenize(value: Any) -> list[str]:
    return [token for token in cell_text(value).split() if token]


def read_external_corpus(workbook_path: Path, labeled_ids: set[str], text_col: str) -> list[list[str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if CORPUS_SHEET not in workbook.sheetnames:
        raise KeyError(f"Sheet not found: {CORPUS_SHEET}")
    sheet = workbook[CORPUS_SHEET]
    headers = [cell.value for cell in sheet[1]]
    if text_col not in headers or ID_COLUMN not in headers:
        raise ValueError(f"{CORPUS_SHEET} must contain {ID_COLUMN} and {text_col}")
    text_idx = headers.index(text_col)
    id_idx = headers.index(ID_COLUMN)

    sentences: list[list[str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        review_id = cell_text(row[id_idx])
        if review_id in labeled_ids:
            continue
        tokens = tokenize(row[text_idx])
        if tokens:
            sentences.append(tokens)
    workbook.close()
    return sentences


def mean_vector(model: Any, value: Any, vector_size: int, allow_oov: bool = False) -> np.ndarray:
    tokens = tokenize(value)
    vectors = []
    for token in tokens:
        if allow_oov:
            vectors.append(model.wv[token])
        elif token in model.wv:
            vectors.append(model.wv[token])
    if not vectors:
        return np.zeros(vector_size, dtype=np.float32)
    return np.mean(vectors, axis=0).astype(np.float32)


def bert_cache_matches(rows: list[dict[str, Any]]) -> bool:
    if not BERT_EMBEDDING_PATH.exists() or not BERT_IDS_PATH.exists():
        return False
    expected_ids = [cell_text(row.get(ID_COLUMN)) for row in rows]
    try:
        cached_ids = json.loads(BERT_IDS_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return False
    if cached_ids != expected_ids:
        return False
    embeddings = np.load(BERT_EMBEDDING_PATH, mmap_mode="r")
    return embeddings.shape == (len(rows), 768)


def build_bert_embeddings(rows: list[dict[str, Any]], text_col: str) -> np.ndarray:
    if bert_cache_matches(rows):
        print(f"BERT embedding cache used: {BERT_EMBEDDING_PATH.name}")
        return np.load(BERT_EMBEDDING_PATH)

    import torch
    from transformers import AutoModel, AutoTokenizer

    print(f"Loading BERT model: {BERT_MODEL_NAME}")
    tokenizer = AutoTokenizer.from_pretrained(BERT_MODEL_NAME)
    model = AutoModel.from_pretrained(BERT_MODEL_NAME)
    model.eval()

    texts = [cell_text(row.get(text_col)) for row in rows]
    embeddings: list[np.ndarray] = []
    with torch.no_grad():
        for start in range(0, len(texts), BERT_BATCH_SIZE):
            batch_texts = texts[start:start + BERT_BATCH_SIZE]
            encoded = tokenizer(
                batch_texts,
                padding=True,
                truncation=True,
                max_length=BERT_MAX_LENGTH,
                return_tensors="pt",
            )
            output = model(**encoded)
            mask = encoded["attention_mask"].unsqueeze(-1).expand(output.last_hidden_state.size()).float()
            pooled = (output.last_hidden_state * mask).sum(dim=1) / mask.sum(dim=1).clamp(min=1e-9)
            embeddings.append(pooled.cpu().numpy().astype(np.float32))
            if (start // BERT_BATCH_SIZE + 1) % 10 == 0:
                print(f"  embedded {min(start + BERT_BATCH_SIZE, len(texts))}/{len(texts)}")

    matrix = np.vstack(embeddings)
    BERT_CACHE_DIR.mkdir(exist_ok=True)
    np.save(BERT_EMBEDDING_PATH, matrix)
    BERT_IDS_PATH.write_text(
        json.dumps([cell_text(row.get(ID_COLUMN)) for row in rows], ensure_ascii=False),
        encoding="utf-8",
    )
    print(f"BERT embeddings saved: {BERT_EMBEDDING_PATH}")
    return matrix


def prepare_resources(
    rows: list[dict[str, Any]],
    workbook_path: Path,
    text_col: str,
    required_vectorizers: set[str] | None = None,
) -> dict[str, Any]:
    resources: dict[str, Any] = {}
    required = required_vectorizers or {config.vectorizer for config in BASELINE_METHODS}
    if {"word2vec", "fasttext"} & required:
        labeled_ids = {cell_text(row.get(ID_COLUMN)) for row in rows}
        corpus = read_external_corpus(workbook_path, labeled_ids, text_col)
        resources["external_corpus"] = corpus
        print(f"External corpus sentences: {len(corpus)}")

        from gensim.models import FastText, Word2Vec

        if "word2vec" in required:
            print("Training Word2Vec mean-pooling model")
            w2v = Word2Vec(
                sentences=corpus,
                vector_size=W2V_VECTOR_SIZE,
                window=W2V_WINDOW,
                min_count=W2V_MIN_COUNT,
                workers=1,
                sg=1,
                epochs=W2V_EPOCHS,
                seed=RANDOM_STATE,
            )
            resources["word2vec"] = {
                cell_text(row.get(ID_COLUMN)): mean_vector(w2v, row.get(text_col), W2V_VECTOR_SIZE)
                for row in rows
            }
            print(f"Word2Vec vocabulary size: {len(w2v.wv)}")

        if "fasttext" in required:
            print("Training FastText mean-pooling model")
            fasttext = FastText(
                sentences=corpus,
                vector_size=FASTTEXT_VECTOR_SIZE,
                window=FASTTEXT_WINDOW,
                min_count=FASTTEXT_MIN_COUNT,
                workers=1,
                sg=1,
                epochs=FASTTEXT_EPOCHS,
                seed=RANDOM_STATE,
            )
            resources["fasttext"] = {
                cell_text(row.get(ID_COLUMN)): mean_vector(fasttext, row.get(text_col), FASTTEXT_VECTOR_SIZE, allow_oov=True)
                for row in rows
            }
            print(f"FastText vocabulary size: {len(fasttext.wv)}")

    if "bert" in required:
        bert_matrix = build_bert_embeddings(rows, text_col)
        resources["bert_ids"] = [cell_text(row.get(ID_COLUMN)) for row in rows]
        resources["bert_index"] = {review_id: index for index, review_id in enumerate(resources["bert_ids"])}
        resources["bert"] = bert_matrix
    return resources


# ==================== 9. MODEL KURULUMU ====================
def class_weight(value: Any) -> Any:
    return None if value in {None, "none", "None"} else value


def make_classifier(config: MethodConfig) -> Any:
    p = config.params
    if config.model == "logistic_regression":
        return LogisticRegression(
            C=float(p["C"]),
            class_weight=class_weight(p.get("class_weight")),
            solver="liblinear",
            max_iter=2000,
            random_state=RANDOM_STATE,
        )
    if config.model == "linear_svm":
        return SVC(
            kernel=str(p.get("kernel", "linear")),
            C=float(p["C"]),
            class_weight=class_weight(p.get("class_weight")),
            gamma=str(p.get("gamma", "scale")),
            degree=int(p.get("degree", 3)),
            random_state=RANDOM_STATE,
            cache_size=500,
        )
    if config.model == "ridge_classifier":
        return RidgeClassifier(
            alpha=float(p["alpha"]),
            class_weight=class_weight(p.get("class_weight")),
        )
    if config.model == "naive_bayes":
        if config.vectorizer in {"word2vec", "fasttext", "bert"}:
            return GaussianNB()
        return MultinomialNB(alpha=float(p["alpha"]))
    if config.model == "knn":
        return KNeighborsClassifier(
            n_neighbors=int(p["n_neighbors"]),
            weights=str(p["weights"]),
            metric=str(p["metric"]),
            algorithm="brute",
            n_jobs=1,
        )
    raise ValueError(f"Unknown model: {config.model}")


def make_word_vectorizer(config: MethodConfig) -> Any:
    p = config.params
    if config.vectorizer == "tfidf":
        return TfidfVectorizer(
            lowercase=False,
            ngram_range=p["ngram_range"],
            min_df=p["min_df"],
            max_df=p["max_df"],
        )
    if config.vectorizer == "count":
        return CountVectorizer(
            lowercase=False,
            ngram_range=p["ngram_range"],
            min_df=p["min_df"],
            max_df=p["max_df"],
        )
    raise ValueError(f"Unknown word vectorizer: {config.vectorizer}")


def make_word_char_vectorizer(config: MethodConfig) -> FeatureUnion:
    p = config.params
    return FeatureUnion([
        ("word", TfidfVectorizer(
            lowercase=False,
            analyzer="word",
            ngram_range=p["word_ngram_range"],
            min_df=p["word_min_df"],
            max_df=p["max_df"],
        )),
        ("char", TfidfVectorizer(
            lowercase=False,
            analyzer="char_wb",
            ngram_range=p["char_ngram_range"],
            min_df=p["char_min_df"],
            max_df=p["max_df"],
        )),
    ])


def make_bert_transformer(name: str) -> Any:
    if name == "l2_normalize":
        return Normalizer(norm="l2")
    if name == "standard_scale":
        return StandardScaler()
    if name == "none":
        return FunctionTransformer(validate=False)
    raise ValueError(f"Unknown BERT transform: {name}")


def make_dense_transformer(name: str) -> Any:
    if name == "l2_normalize":
        return Normalizer(norm="l2")
    if name == "standard_scale":
        return StandardScaler()
    if name == "none":
        return FunctionTransformer(validate=False)
    raise ValueError(f"Unknown dense transform: {name}")


def make_estimator(config: MethodConfig) -> Pipeline:
    p = config.params
    if config.vectorizer in {"tfidf", "count"}:
        return Pipeline([("vectorizer", make_word_vectorizer(config)), ("clf", make_classifier(config))])
    if config.vectorizer == "tfidf_word_char":
        return Pipeline([("vectorizer", make_word_char_vectorizer(config)), ("clf", make_classifier(config))])
    if config.vectorizer in {"word2vec", "fasttext"}:
        if config.model == "knn":
            return Pipeline([
                ("prep", make_dense_transformer(str(p.get("transform", "none")))),
                ("clf", make_classifier(config)),
            ])
        return Pipeline([("clf", make_classifier(config))])
    if config.vectorizer == "bert":
        if config.model == "knn":
            return Pipeline([
                ("prep", make_bert_transformer(str(p.get("transform", "standard_scale")))),
                ("clf", make_classifier(config)),
            ])
        if config.model == "naive_bayes":
            return Pipeline([("clf", make_classifier(config))])
        return Pipeline([("scale", StandardScaler()), ("clf", make_classifier(config))])
    raise ValueError(f"Unknown vectorizer: {config.vectorizer}")


# ==================== 10. X/Y HAZIRLAMA ====================
def features_for_rows(
    rows: list[dict[str, Any]],
    config: MethodConfig,
    resources: dict[str, Any],
    text_col: str,
) -> Any:
    if config.vectorizer in {"tfidf", "count", "tfidf_word_char"}:
        return [cell_text(row.get(text_col)) for row in rows]
    if config.vectorizer in {"word2vec", "fasttext"}:
        vector_map = resources[config.vectorizer]
        return np.vstack([vector_map[cell_text(row.get(ID_COLUMN))] for row in rows])
    if config.vectorizer == "bert":
        bert = resources["bert"]
        index = resources["bert_index"]
        return np.vstack([bert[index[cell_text(row.get(ID_COLUMN))]] for row in rows])
    raise ValueError(f"Unknown vectorizer: {config.vectorizer}")


def train_features_labels(
    rows: list[dict[str, Any]],
    config: MethodConfig,
    resources: dict[str, Any],
    text_col: str,
) -> tuple[Any, list[str], list[int]]:
    tr = train_rows(rows)
    x = features_for_rows(tr, config, resources, text_col)
    y = to_y(tr)
    folds = [int(cell_text(row.get(FOLD_COLUMN))) for row in tr]
    return x, y, folds


def final_features_labels(
    rows: list[dict[str, Any]],
    config: MethodConfig,
    resources: dict[str, Any],
    text_col: str,
) -> tuple[Any, list[str], Any, list[str]]:
    tr = train_rows(rows)
    te = test_rows(rows)
    x_train = features_for_rows(tr, config, resources, text_col)
    y_train = to_y(tr)
    x_test = features_for_rows(te, config, resources, text_col)
    y_test = to_y(te)
    return x_train, y_train, x_test, y_test


# ==================== 11. 10-FOLD VE TEST ====================
def evaluate_estimator_10fold(
    x: Any,
    y: list[str],
    folds: list[int],
    estimator: Any,
) -> dict[str, Any]:
    fold_rows: list[dict[str, Any]] = []
    y_array = np.array(y)
    fold_array = np.array(folds)
    for fold in range(1, 11):
        train_mask = fold_array != fold
        valid_mask = fold_array == fold
        model = clone(estimator)
        if isinstance(x, list):
            x_train = [x[i] for i in np.where(train_mask)[0]]
            x_valid = [x[i] for i in np.where(valid_mask)[0]]
        else:
            x_train = x[train_mask]
            x_valid = x[valid_mask]
        y_train = list(y_array[train_mask])
        y_valid = list(y_array[valid_mask])
        model.fit(x_train, y_train)
        fold_metric = metrics(y_valid, list(model.predict(x_valid)))
        fold_metric["fold"] = fold
        fold_rows.append(fold_metric)
    return mean_metrics(fold_rows)


def evaluate_estimator_final(
    x_train: Any,
    y_train: list[str],
    x_test: Any,
    y_test: list[str],
    estimator: Any,
) -> dict[str, Any]:
    model = clone(estimator)
    model.fit(x_train, y_train)
    return metrics(y_test, list(model.predict(x_test)))


def evaluate_config(
    rows: list[dict[str, Any]],
    config: MethodConfig,
    resources: dict[str, Any],
    text_col: str,
) -> dict[str, Any]:
    estimator = make_estimator(config)
    x_train_data, y_train_data, folds = train_features_labels(rows, config, resources, text_col)
    fold = evaluate_estimator_10fold(x_train_data, y_train_data, folds, estimator)
    x_train, y_train, x_test, y_test = final_features_labels(rows, config, resources, text_col)
    final = evaluate_estimator_final(x_train, y_train, x_test, y_test, estimator)
    return {
        "name": config.name,
        "vectorizer": vectorizer_label(config),
        "model": model_label(config),
        "representation_params": representation_param_summary(config),
        "model_params": model_param_summary(config),
        "params": param_summary(config),
        **fold,
        **{f"final_{key}": value for key, value in final.items()},
    }


# ==================== 12. GRID SEARCH ====================
def grid_for_method(config: MethodConfig) -> dict[str, list[Any]]:
    if config.vectorizer in {"tfidf", "count"}:
        grid: dict[str, list[Any]] = {
            "vectorizer__ngram_range": [config.params["ngram_range"]],
            "vectorizer__min_df": [config.params["min_df"]],
            "vectorizer__max_df": [config.params["max_df"]],
        }
    elif config.vectorizer == "tfidf_word_char":
        grid = {
            "vectorizer__word__ngram_range": [config.params["word_ngram_range"]],
            "vectorizer__word__min_df": [config.params["word_min_df"]],
            "vectorizer__word__max_df": [config.params["max_df"]],
            "vectorizer__char__ngram_range": [config.params["char_ngram_range"]],
            "vectorizer__char__min_df": [config.params["char_min_df"]],
            "vectorizer__char__max_df": [config.params["max_df"]],
        }
    elif config.vectorizer in {"word2vec", "fasttext", "bert"}:
        grid = {}
    else:
        raise ValueError(f"Unknown vectorizer: {config.vectorizer}")

    if config.model == "knn":
        grid.update({
            "clf__n_neighbors": K_VALUES,
            "clf__weights": KNN_WEIGHTS,
        })
        if config.vectorizer in {"word2vec", "fasttext", "bert"}:
            grid.update({"prep": [FunctionTransformer(validate=False), Normalizer(norm="l2"), StandardScaler()]})
        return grid
    if config.model == "logistic_regression":
        grid.update({"clf__C": LOGISTIC_C_VALUES, "clf__class_weight": CLASS_WEIGHT_VALUES})
    elif config.model == "linear_svm":
        grid.update({
            "clf__kernel": SVM_KERNEL_VALUES,
            "clf__C": SVM_C_VALUES,
            "clf__class_weight": CLASS_WEIGHT_VALUES,
            "clf__gamma": SVM_GAMMA_VALUES,
            "clf__degree": SVM_DEGREE_VALUES,
        })
    elif config.model == "ridge_classifier":
        grid.update({"clf__alpha": RIDGE_ALPHA_VALUES, "clf__class_weight": CLASS_WEIGHT_VALUES})
    elif config.model == "naive_bayes":
        if config.vectorizer in {"word2vec", "fasttext", "bert"}:
            return grid
        grid.update({"clf__alpha": NB_ALPHA_VALUES})
    else:
        raise ValueError(f"Unknown model: {config.model}")
    return grid


def format_best_params(best_params: dict[str, Any]) -> str:
    parts = []
    for key, value in best_params.items():
        clean_key = key.replace("vectorizer__", "").replace("clf__", "").replace("prep", "transform")
        if isinstance(value, FunctionTransformer):
            value = "none"
        elif isinstance(value, Normalizer):
            value = "l2_normalize"
        elif isinstance(value, StandardScaler):
            value = "standard_scale"
        parts.append(f"{clean_key}={value}")
    return ", ".join(parts)


def run_grid_search(
    rows: list[dict[str, Any]],
    configs: list[MethodConfig],
    resources: dict[str, Any],
    text_col: str,
) -> list[dict[str, Any]]:
    outputs: list[dict[str, Any]] = []
    for config in configs:
        print(f"Grid search: {config.name}")
        x_train_data, y_train_data, folds = train_features_labels(rows, config, resources, text_col)
        cv = PredefinedSplit([fold - 1 for fold in folds])
        search = GridSearchCV(
            estimator=make_estimator(config),
            param_grid=grid_for_method(config),
            scoring=make_scorer(f1_score, pos_label="1", zero_division=0),
            cv=cv,
            n_jobs=1,
            refit=True,
            error_score="raise",
        )
        search.fit(x_train_data, y_train_data)
        best_estimator = search.best_estimator_
        fold = evaluate_estimator_10fold(x_train_data, y_train_data, folds, best_estimator)
        x_train, y_train, x_test, y_test = final_features_labels(rows, config, resources, text_col)
        final = evaluate_estimator_final(x_train, y_train, x_test, y_test, best_estimator)
        best_params = format_best_params(search.best_params_) or model_param_summary(config)
        outputs.append({
            "name": config.name,
            "vectorizer": vectorizer_label(config),
            "model": model_label(config),
            "representation_params": representation_param_summary(config),
            "model_params": best_params,
            "params": best_params,
            "grid_best_f1": search.best_score_,
            **fold,
            **{f"final_{key}": value for key, value in final.items()},
        })
        print(
            f"  best_f1={search.best_score_:.4f} "
            f"final_acc={final['accuracy']:.4f} final_f1={final['f1_score']:.4f} "
            f"dogru={final['correct']}/{final['total']}"
        )
    return outputs


# ==================== 13. RAPOR METINLERI ====================
def vectorizer_label(config: MethodConfig) -> str:
    labels = {
        "tfidf": "TF-IDF",
        "count": "BoW",
        "tfidf_word_char": "TF-IDF + Character N-gram",
        "word2vec": "Word2Vec",
        "fasttext": "FastText",
        "bert": "BERT",
    }
    return labels[config.vectorizer]


def model_label(config: MethodConfig) -> str:
    labels = {
        "logistic_regression": "Logistic Regression",
        "linear_svm": "SVM",
        "ridge_classifier": "Ridge Classifier",
        "naive_bayes": "Naive Bayes",
        "knn": "KNN",
    }
    return labels[config.model]


def param_summary(config: MethodConfig) -> str:
    return ", ".join(f"{key}={value}" for key, value in config.params.items())


def representation_param_summary(config: MethodConfig) -> str:
    p = config.params
    if config.vectorizer in {"tfidf", "count"}:
        return f"word_ngram={p['ngram_range']}; min_df={p['min_df']}; max_df={p['max_df']}"
    if config.vectorizer == "tfidf_word_char":
        return (
            f"word_ngram={p['word_ngram_range']}; char_ngram={p['char_ngram_range']}; "
            f"word_min_df={p['word_min_df']}; char_min_df={p['char_min_df']}; max_df={p['max_df']}"
        )
    if config.vectorizer in {"word2vec", "fasttext"}:
        return f"mean pooling; vector_size={W2V_VECTOR_SIZE if config.vectorizer == 'word2vec' else FASTTEXT_VECTOR_SIZE}; corpus=unlabeled Goodreads reviews"
    if config.vectorizer == "bert":
        return f"mean pooling; model={BERT_MODEL_NAME}; max_length={p.get('max_length', BERT_MAX_LENGTH)}"
    return param_summary(config)


def model_param_summary(config: MethodConfig) -> str:
    p = config.params
    if config.model == "logistic_regression":
        return f"C={p['C']}; class_weight={p.get('class_weight')}"
    if config.model == "linear_svm":
        return f"kernel={p.get('kernel', 'linear')}; C={p['C']}; class_weight={p.get('class_weight')}"
    if config.model == "ridge_classifier":
        return f"alpha={p['alpha']}; class_weight={p.get('class_weight')}"
    if config.model == "naive_bayes":
        if config.vectorizer in {"word2vec", "fasttext", "bert"}:
            return "variant=GaussianNB"
        return f"variant=MultinomialNB; alpha={p['alpha']}"
    if config.model == "knn":
        transform = p.get("transform")
        prefix = f"transform={transform}; " if transform else ""
        return f"{prefix}n_neighbors={p['n_neighbors']}; weights={p['weights']}; metric={p['metric']}"
    return param_summary(config)


def top_configs_from_baseline(baseline_rows: list[dict[str, Any]], top_n: int) -> list[MethodConfig]:
    order = sorted(
        baseline_rows,
        key=lambda row: (float(row["fold_f1_score"]), float(row["fold_accuracy"]), int(row["fold_correct"])),
        reverse=True,
    )
    names = [row["name"] for row in order[:top_n]]
    return [config for name in names for config in BASELINE_METHODS if config.name == name]


def read_baseline_rows_from_sheet(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if RESULTS_SHEET not in workbook.sheetnames:
        raise KeyError(f"Sheet not found: {RESULTS_SHEET}")
    sheet = workbook[RESULTS_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    start_index = None
    for index, row in enumerate(rows):
        if row and row[0] in {"Fixed Method Results", "Sabit Yontem Sonuclari"}:
            start_index = index + 2
            break
    if start_index is None:
        raise ValueError("Sabit Yontem Sonuclari tablosu bulunamadi.")

    header = list(rows[start_index - 1])
    output: list[dict[str, Any]] = []
    for row in rows[start_index:]:
        if not row or row[0] in {None, "", "Grid Search Top 5 Sonuclari"}:
            break
        data = {header[col]: row[col] if col < len(row) else None for col in range(len(header)) if header[col]}
        output.append({
            "name": data.get("Method") or f"{data.get('Classification Method')} + {data.get('Text Representation')}",
            "vectorizer": data.get("Text Representation") or data.get("Vektorlestirme"),
            "model": data.get("Classification Method") or data.get("Model"),
            "representation_params": data.get("Representation Parameters") or data.get("Temsil ayari") or "",
            "model_params": data.get("Classification Parameters") or data.get("Model ayari") or data.get("Parametre") or "",
            "params": data.get("Classification Parameters") or data.get("Parametre") or "",
            "fold_accuracy": data["10-Fold Accuracy"],
            "fold_f1_score": data.get("10-Fold F1") or data.get("10-Fold F1-score"),
            "fold_precision": data["10-Fold Precision"],
            "fold_recall": data["10-Fold Recall"],
            "fold_f1_score_std": data["10-Fold F1 Std"],
            "fold_correct": data.get("10-Fold Correct") or data.get("10-Fold Dogru"),
            "fold_total": data.get("10-Fold Total") or data.get("10-Fold Toplam"),
            "final_accuracy": data["Final Accuracy"],
            "final_f1_score": data.get("Final F1") or data.get("Final F1-score"),
            "final_precision": data["Final Precision"],
            "final_recall": data["Final Recall"],
            "final_correct": data.get("Final Correct") or data.get("Final Dogru"),
            "final_total": data.get("Final Total") or data.get("Final Toplam"),
            "final_tn": data["TN"],
            "final_fp": data["FP"],
            "final_fn": data["FN"],
            "final_tp": data["TP"],
        })
    if len(output) != len(BASELINE_METHODS):
        raise ValueError(f"Expected {len(BASELINE_METHODS)} baseline rows, found {len(output)}")
    return output


def read_grid_rows_from_sheet(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if RESULTS_SHEET not in workbook.sheetnames:
        workbook.close()
        return []
    sheet = workbook[RESULTS_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    start_index = None
    for index, row in enumerate(rows):
        if row and row[0] in {"GridSearchCV Results", "GridSearchCV Top 5 Results", "Grid Search Top 5 Sonuclari"}:
            start_index = index + 2
            break
    if start_index is None:
        return []

    header = list(rows[start_index - 1])
    output: list[dict[str, Any]] = []
    for row in rows[start_index:]:
        if not row or not row[0]:
            break
        data = {header[col]: row[col] if col < len(row) else None for col in range(len(header)) if header[col]}
        if data.get("Grid Best F1") in {None, "", "nan"}:
            continue
        output.append({
            "name": data.get("Method") or f"{data.get('Classification Method')} + {data.get('Text Representation')}",
            "vectorizer": data.get("Text Representation") or data.get("Vektorlestirme"),
            "model": data.get("Classification Method") or data.get("Model"),
            "representation_params": data.get("Representation Parameters") or data.get("Temsil ayari") or "",
            "model_params": data.get("Classification Parameters") or data.get("Model ayari") or data.get("Parametre") or "",
            "params": data.get("Classification Parameters") or data.get("Parametre") or "",
            "grid_best_f1": data["Grid Best F1"],
            "fold_accuracy": data["10-Fold Accuracy"],
            "fold_f1_score": data.get("10-Fold F1") or data.get("10-Fold F1-score"),
            "fold_precision": data["10-Fold Precision"],
            "fold_recall": data["10-Fold Recall"],
            "fold_f1_score_std": data["10-Fold F1 Std"],
            "fold_correct": data.get("10-Fold Correct") or data.get("10-Fold Dogru"),
            "fold_total": data.get("10-Fold Total") or data.get("10-Fold Toplam"),
            "final_accuracy": data["Final Accuracy"],
            "final_f1_score": data.get("Final F1") or data.get("Final F1-score"),
            "final_precision": data["Final Precision"],
            "final_recall": data["Final Recall"],
            "final_correct": data.get("Final Correct") or data.get("Final Dogru"),
            "final_total": data.get("Final Total") or data.get("Final Toplam"),
            "final_tn": data["TN"],
            "final_fp": data["FP"],
            "final_fn": data["FN"],
            "final_tp": data["TP"],
        })
    return output


def merge_grid_rows(existing_rows: list[dict[str, Any]], new_rows: list[dict[str, Any]], order: list[MethodConfig]) -> list[dict[str, Any]]:
    merged = {row["name"]: row for row in existing_rows}
    for row in new_rows:
        merged[row["name"]] = row
    ordered_names = [config.name for config in order]
    return [merged[name] for name in ordered_names if name in merged]


# ==================== 14. EXCEL YAZMA ====================
def append_result_table(sheet: Any, title: str, rows: list[dict[str, Any]], include_grid_score: bool = False) -> None:
    sheet.append([])
    sheet.append([title])
    header = [
        "Classification Method", "Text Representation", "Representation Parameters", "Classification Parameters",
    ]
    if include_grid_score:
        header.append("Grid Best F1")
    header += [
        "10-Fold Accuracy", "10-Fold F1", "10-Fold Precision", "10-Fold Recall",
        "10-Fold F1 Std", "10-Fold Correct", "10-Fold Total",
        "Final Accuracy", "Final F1", "Final Precision", "Final Recall",
        "Final Correct", "Final Total", "TN", "FP", "FN", "TP",
    ]
    sheet.append(header)
    for row in rows:
        values = [
            row["model"], row["vectorizer"], row.get("representation_params", ""), row.get("model_params", row.get("params", "")),
        ]
        if include_grid_score:
            values.append(round(float(row["grid_best_f1"]), 4))
        values += [
            round(float(row["fold_accuracy"]), 4), round(float(row["fold_f1_score"]), 4),
            round(float(row["fold_precision"]), 4), round(float(row["fold_recall"]), 4),
            round(float(row["fold_f1_score_std"]), 4), row["fold_correct"], row["fold_total"],
            round(float(row["final_accuracy"]), 4), round(float(row["final_f1_score"]), 4),
            round(float(row["final_precision"]), 4), round(float(row["final_recall"]), 4),
            row["final_correct"], row["final_total"],
            row["final_tn"], row["final_fp"], row["final_fn"], row["final_tp"],
        ]
        sheet.append(values)


def write_model_sheet(
    workbook_path: Path,
    baseline_rows: list[dict[str, Any]],
    grid_rows: list[dict[str, Any]] | None = None,
) -> None:
    workbook = load_workbook(workbook_path)
    if RESULTS_SHEET in workbook.sheetnames:
        del workbook[RESULTS_SHEET]
    sheet = workbook.create_sheet(RESULTS_SHEET)
    sheet.append(["Informative Model Comparison - V4"])
    sheet.append(["Scope", "10-fold train validation + test reporting"])
    sheet.append(["Design", "5 classification methods x 6 text representations = 30 fixed combinations."])
    sheet.append(["Grid selection metric", "10-fold train F1"])
    sheet.append(["Reported metrics", "Accuracy, F1, Precision, Recall, Correct/Total"])
    sheet.append([])
    sheet.append(["Classification Method Grid Search Spaces"])
    sheet.append(["Logistic Regression", f"C={LOGISTIC_C_VALUES}; class_weight={CLASS_WEIGHT_VALUES}"])
    sheet.append([
        "SVM",
        f"kernel={SVM_KERNEL_VALUES}; C={SVM_C_VALUES}; class_weight={CLASS_WEIGHT_VALUES}; "
        f"gamma={SVM_GAMMA_VALUES}; degree={SVM_DEGREE_VALUES}",
    ])
    sheet.append(["Ridge Classifier", f"alpha={RIDGE_ALPHA_VALUES}; class_weight={CLASS_WEIGHT_VALUES}"])
    sheet.append([
        "Naive Bayes",
        f"MultinomialNB: alpha={NB_ALPHA_VALUES} for BoW/TF-IDF; "
        "GaussianNB for Word2Vec/FastText/BERT",
    ])
    sheet.append([
        "KNN",
        f"n_neighbors={K_VALUES}; weights={KNN_WEIGHTS}; metric=cosine; "
        f"dense_transform={BERT_KNN_TRANSFORMS}",
    ])

    sheet.append([])
    sheet.append(["Text Representation Fixed Settings"])
    sheet.append(["BoW", "ngram_range=(1, 2); min_df=1; max_df=0.90"])
    sheet.append(["TF-IDF", "word_ngram_range=(1, 1); min_df=2; max_df=0.90"])
    sheet.append([
        "TF-IDF + Character N-gram",
        "word_ngram_range=(1, 1); char_ngram_range=(3, 6); "
        "word_min_df=2; char_min_df=2; max_df=0.90",
    ])
    sheet.append([
        "Word2Vec",
        f"trained on unlabeled cleaned Goodreads reviews; vector_size={W2V_VECTOR_SIZE}; "
        f"window={W2V_WINDOW}; min_count={W2V_MIN_COUNT}; epochs={W2V_EPOCHS}; sg=1; pooling=mean; fixed setting",
    ])
    sheet.append([
        "FastText",
        f"trained on unlabeled cleaned Goodreads reviews; vector_size={FASTTEXT_VECTOR_SIZE}; "
        f"window={FASTTEXT_WINDOW}; min_count={FASTTEXT_MIN_COUNT}; epochs={FASTTEXT_EPOCHS}; sg=1; pooling=mean; fixed setting",
    ])
    sheet.append([
        "BERT",
        f"pre-trained {BERT_MODEL_NAME}; used for embedding extraction; max_length={BERT_MAX_LENGTH}; "
        "pooling=mean; fixed setting",
    ])

    sheet.append([])
    sheet.append(["Fixed Parameter Settings"])
    sheet.append(["Classification Method", "Text Representation", "Representation Parameters", "Classification Parameters"])
    for config in BASELINE_METHODS:
        sheet.append([model_label(config), vectorizer_label(config), representation_param_summary(config), model_param_summary(config)])

    append_result_table(sheet, "Fixed Method Results", baseline_rows)
    if grid_rows:
        grid_title = "GridSearchCV Results" if len(grid_rows) == len(BASELINE_METHODS) else "GridSearchCV Top 5 Results"
        append_result_table(sheet, grid_title, grid_rows, include_grid_score=True)

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    title_fill = PatternFill("solid", fgColor="B6D7A8")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    for row in sheet.iter_rows():
        values = [cell.value for cell in row]
        if values and values[0] in {
            "Common Grid Search Spaces",
            "Classification Method Grid Search Spaces",
            "Text Representation Fixed Settings",
            "Fixed Parameter Settings",
            "Fixed Method Results",
            "GridSearchCV Results",
            "GridSearchCV Top 5 Results",
        }:
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = title_fill
        if values and values[0] == "Classification Method":
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = header_fill

    widths = {
        "A": 45, "B": 32, "C": 22, "D": 72, "E": 16, "F": 16, "G": 18, "H": 16,
        "I": 14, "J": 14, "K": 14, "L": 15, "M": 15, "N": 16, "O": 14,
        "P": 12, "Q": 12, "R": 8, "S": 8, "T": 8, "U": 8, "V": 8,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = "A1"

    if "Informative_Hata_Analizi" in workbook.sheetnames:
        workbook._sheets.remove(sheet)
        target_index = workbook.sheetnames.index("Informative_Hata_Analizi") + 1
        workbook._sheets.insert(target_index, sheet)
        workbook.active = target_index
    else:
        workbook.active = workbook.sheetnames.index(RESULTS_SHEET)
    workbook.save(workbook_path)


# ==================== 15. ANA AKIS ====================
def run_baseline(
    rows: list[dict[str, Any]],
    resources: dict[str, Any],
    text_col: str,
) -> list[dict[str, Any]]:
    outputs: list[dict[str, Any]] = []
    for index, config in enumerate(BASELINE_METHODS, start=1):
        print(f"[{index}/{len(BASELINE_METHODS)}] Calisiyor: {config.name}")
        outputs.append(evaluate_config(rows, config, resources, text_col))
    return outputs


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    rows = read_rows(workbook_path, args.sheet)
    validate_dataset(rows, args.text_col)
    if args.mode == "validate":
        return

    if args.mode == "grid_top5_only":
        baseline_rows = read_baseline_rows_from_sheet(workbook_path)
        top_configs = top_configs_from_baseline(baseline_rows, args.top_n)
        if args.grid_method:
            top_names = {config.name for config in top_configs}
            if args.grid_method not in top_names:
                raise ValueError(f"--grid-method top {args.top_n} icinde degil: {args.grid_method}")
            selected_configs = [config for config in top_configs if config.name == args.grid_method]
        else:
            selected_configs = top_configs
        resources = prepare_resources(
            rows,
            workbook_path,
            args.text_col,
            {config.vectorizer for config in selected_configs},
        )
        print("Grid search methods:", ", ".join(config.name for config in selected_configs))
        new_grid_rows = run_grid_search(rows, selected_configs, resources, args.text_col)
        grid_rows = merge_grid_rows(read_grid_rows_from_sheet(workbook_path), new_grid_rows, top_configs)
        if not args.no_excel:
            write_model_sheet(workbook_path, baseline_rows, grid_rows)
            print(f"Excel sheet written: {RESULTS_SHEET}")
        return

    resources = prepare_resources(rows, workbook_path, args.text_col)
    baseline_rows = run_baseline(rows, resources, args.text_col)
    grid_rows: list[dict[str, Any]] | None = None

    if not args.no_excel:
        write_model_sheet(workbook_path, baseline_rows, None)
        print(f"Baseline Excel sheet written: {RESULTS_SHEET}")

    if args.mode == "grid_top5":
        top_configs = top_configs_from_baseline(baseline_rows, args.top_n)
        print("Grid search top methods:", ", ".join(config.name for config in top_configs))
        grid_rows = run_grid_search(rows, top_configs, resources, args.text_col)
    elif args.mode == "grid_all":
        print("Grid search all methods:", ", ".join(config.name for config in BASELINE_METHODS))
        grid_rows = run_grid_search(rows, BASELINE_METHODS, resources, args.text_col)

    if not args.no_excel:
        write_model_sheet(workbook_path, baseline_rows, grid_rows)
        print(f"Excel sheet written: {RESULTS_SHEET}")


if __name__ == "__main__":
    main()
